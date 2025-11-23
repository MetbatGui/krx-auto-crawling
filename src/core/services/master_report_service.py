"""
마스터 리포트 비즈니스 로직 서비스

전체 워크플로우를 오케스트레이션하고 StoragePort를 통해 파일 I/O를 수행합니다.
"""
import pandas as pd
import datetime
from typing import Dict, List
from pathlib import Path

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

from core.ports.storage_port import StoragePort
from core.domain.models import KrxData


class MasterReportService:
    """마스터 리포트 비즈니스 로직 서비스"""
    
    def __init__(self, storage: StoragePort, file_name_prefix: str = "2025"):
        """
        Args:
            storage: 파일 저장/로드를 위한 StoragePort
            file_name_prefix: 파일명에 사용될 연도 접두사
        """
        self.storage = storage
        self.excel_columns = ['일자', '종목', '금액']
        
        # 파일 경로 설정
        self.master_subdir = "순매수도"
        year_suffix = f"({file_name_prefix})"
        self.file_map: Dict[str, str] = {
            'KOSPI_foreigner': f'코스피외국인순매수도{year_suffix}.xlsx',
            'KOSDAQ_foreigner': f'코스닥외국인순매수도{year_suffix}.xlsx',
            'KOSPI_institutions': f'코스피기관순매수도{year_suffix}.xlsx',
            'KOSDAQ_institutions': f'코스닥기관순매수도{year_suffix}.xlsx',
        }
        
        # 순매수도 디렉토리 생성
        self.storage.ensure_directory(self.master_subdir)
    
    def transform_to_excel_schema(
        self,
        daily_data: pd.DataFrame,
        date_int: int
    ) -> pd.DataFrame:
        """
        일별 데이터를 Excel 스키마로 변환합니다.
        
        Args:
            daily_data: KRX 일별 데이터 (종목명, 순매수_거래대금 컬럼 포함)
            date_int: 날짜 정수 (예: 20251121)
            
        Returns:
            변환된 DataFrame (일자, 종목, 금액 컬럼)
        """
        try:
            data_dict = {
                '일자': date_int,
                '종목': daily_data['종목명'],
                '금액': pd.to_numeric(daily_data['순매수_거래대금'])
            }
            
            formatted_df = pd.DataFrame(data_dict)
            formatted_df = formatted_df[self.excel_columns]
            
            print(f"    -> [Service:MasterReport] 데이터 변환 완료 ({len(formatted_df)}개 종목)")
            return formatted_df
            
        except Exception as e:
            print(f"    -> [Service:MasterReport] 🚨 데이터 변환 실패: {e}")
            raise
    
    def check_duplicate_date(
        self,
        existing_df: pd.DataFrame,
        date_int: int
    ) -> bool:
        """
        중복 날짜가 있는지 확인합니다.
        
        Args:
            existing_df: 기존 데이터 DataFrame
            date_int: 확인할 날짜 정수
            
        Returns:
            True if 중복 존재, False otherwise
        """
        if existing_df.empty:
            return False
        
        is_duplicate = date_int in existing_df['일자'].values
        
        if is_duplicate:
            print(f"    -> [Service:MasterReport] ⚠️ {date_int} 데이터 중복 발견")
        
        return is_duplicate
    
    def merge_data(
        self,
        existing_df: pd.DataFrame,
        new_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        기존 데이터와 신규 데이터를 병합합니다.
        
        Args:
            existing_df: 기존 데이터
            new_df: 신규 데이터
            
        Returns:
            병합된 DataFrame
        """
        if existing_df.empty:
            merged = new_df.copy()
        else:
            merged = pd.concat([existing_df, new_df], ignore_index=True)
        
        print(f"    -> [Service:MasterReport] 데이터 병합 완료 (총 {len(merged)}줄)")
        return merged
    
    def create_empty_dataframe(self) -> pd.DataFrame:
        """
        빈 Excel 스키마 DataFrame을 생성합니다.
        
        Returns:
            빈 DataFrame (일자, 종목, 금액 컬럼)
        """
        return pd.DataFrame(columns=self.excel_columns)
    
    def _calculate_pivot(
        self, 
        data: pd.DataFrame, 
        date_int: int
    ) -> pd.DataFrame:
        """
        피벗 테이블을 계산합니다.
        
        Args:
            data: 원본 데이터 (일자, 종목, 금액 컬럼 포함)
            date_int: 기준 날짜 (피벗 컬럼에서 찾기 위함)
            
        Returns:
            정렬된 피벗 DataFrame (총계 포함)
        """
        if data.empty:
            print(f"    -> [Service:MasterReport] ⚠️ 데이터가 비어있어 피벗을 생성할 수 없습니다.")
            return pd.DataFrame()
        
        try:
            # 1. 금액 컬럼 정제 (문자열 -> 숫자)
            data = data.copy()
            data['금액'] = data['금액'].astype(str).str.replace(r'[^0-9.-]', '', regex=True)
            data['금액'] = data['금액'].replace('', 0)
            data['금액'] = pd.to_numeric(data['금액'], errors='coerce').fillna(0)
            
            # 2. 피벗 테이블 생성
            pivot = pd.pivot_table(
                data,
                values='금액',
                index='종목',
                columns='일자',
                aggfunc='sum'
            )
            
            # 3. 총계 추가 및 정렬
            pivot['총계'] = pivot.sum(axis=1)
            pivot_sorted = pivot.sort_values(by='총계', ascending=False)
            
            print(f"    -> [Service:MasterReport] 피벗 테이블 계산 완료")
            return pivot_sorted
            
        except Exception as e:
            print(f"    -> [Service:MasterReport] 🚨 피벗 계산 실패: {e}")
            return pd.DataFrame()
    
    def _should_skip(self, file_path: str, pivot_sheet_name: str) -> bool:
        """
        빠른 건너뛰기 확인 - 피벗 시트가 이미 존재하는지 체크합니다.
        
        Args:
            file_path: 파일 경로 (상대 경로)
            pivot_sheet_name: 피벗 시트 이름
            
        Returns:
            건너뛰기 여부 (True면 이미 처리됨)
        """
        if not self.storage.path_exists(file_path):
            return False
            
        try:
            book = self.storage.load_workbook(file_path)
            if book and pivot_sheet_name in book.sheetnames:
                book.close()
                print(f"    -> [Service:MasterReport] ⚠️ '{pivot_sheet_name}' 피벗 시트 존재 - 건너뛰기")
                return True
        except Exception as e:
            print(f"    -> [Service:MasterReport] ⚠️ 건너뛰기 확인 중 오류: {e}")
        
        return False
    
    def _load_existing_data(
        self, 
        file_path: str, 
        sheet_name: str
    ) -> pd.DataFrame:
        """
        기존 엑셀 데이터를 로드합니다.
        
        Args:
            file_path: 파일 경로 (상대 경로)
            sheet_name: 시트 이름
            
        Returns:
            로드된 DataFrame (파일이 없으면 빈 DataFrame)
        """
        if not self.storage.path_exists(file_path):
            print(f"    -> [Service:MasterReport] 새 파일이 생성됩니다")
            return pd.DataFrame(columns=self.excel_columns)
            
        try:
            # StoragePort의 base_path를 고려하여 전체 경로 구성
            full_path = Path(self.storage.base_path) / file_path
            
            df = pd.read_excel(
                full_path,
                sheet_name=sheet_name,
                engine='openpyxl',
                skiprows=1,
                dtype={'일자': int}
            )
            
            if not df.empty and all(col in df.columns for col in self.excel_columns):
                result = df[self.excel_columns].copy()
                print(f"    -> [Service:MasterReport] 기존 '{sheet_name}' 시트 데이터 ({len(result)}줄) 로드 완료")
                return result
            else:
                print(f"    -> [Service:MasterReport] ⚠️ {sheet_name} 시트 헤더가 손상됨")
                return pd.DataFrame(columns=self.excel_columns)
                
        except (FileNotFoundError, ValueError, KeyError) as e:
            print(f"    -> [Service:MasterReport] ⚠️ 시트가 없어 새로 생성합니다")
            return pd.DataFrame(columns=self.excel_columns)
        except Exception as e:
            print(f"    -> [Service:MasterReport] 🚨 파일 로드 실패: {e}")
            raise
    
    def _save_workbook(
        self,
        file_path: str,
        sheet_name: str,
        pivot_sheet_name: str,
        new_data: pd.DataFrame,
        pivot_data: pd.DataFrame,
        date_int: int,
        sheet_exists: bool
    ) -> bool:
        """
        Excel 워크북을 생성하고 저장합니다.
        
        Args:
            file_path: 파일 경로 (상대 경로)
            sheet_name: Raw 데이터 시트 이름
            pivot_sheet_name: 피벗 시트 이름
            new_data: 추가할 새 데이터
            pivot_data: 피벗 테이블 데이터
            date_int: 기준 날짜
            sheet_exists: Raw 시트 존재 여부
            
        Returns:
            저장 성공 여부
        """
        try:
            # 1. 워크북 로드 또는 생성
            book = self.storage.load_workbook(file_path)
            if book is None:
                book = openpyxl.Workbook()
                if 'Sheet' in book.sheetnames:
                    book.remove(book['Sheet'])
            
            # 2. Raw 데이터 시트 업데이트
            if not new_data.empty:
                self._update_raw_sheet(book, sheet_name, new_data, sheet_exists)
            
            # 3. 피벗 시트 생성
            self._create_pivot_sheet(
                book, sheet_name, pivot_sheet_name,
                pivot_data, date_int
            )
            
            # 4. 저장
            success = self.storage.save_workbook(book, file_path)
            if success:
                print(f"    -> [Service:MasterReport] ✅ Excel 파일 저장 완료")
                if not pivot_data.empty:
                    print(f"    -> [Service:MasterReport] 피벗 샘플:\n{pivot_data.head()}")
            return success
            
        except Exception as e:
            print(f"    -> [Service:MasterReport] 🚨 워크북 저장 실패: {e}")
            return False
    
    def _update_raw_sheet(
        self,
        book: openpyxl.Workbook,
        sheet_name: str,
        new_data: pd.DataFrame,
        sheet_exists: bool
    ) -> None:
        """Raw 데이터 시트를 업데이트합니다."""
        if sheet_exists and sheet_name in book.sheetnames:
            # 기존 시트에 추가
            ws = book[sheet_name]
            print(f"    -> [Service:MasterReport] '{sheet_name}' 시트에 데이터 추가")
            for row in dataframe_to_rows(new_data, index=False, header=False):
                ws.append(row)
        else:
            # 새 시트 생성 (마지막 시트 앞에)
            data_sheet_index = max(0, len(book.sheetnames) - 1) if book.sheetnames else 0
            ws = book.create_sheet(title=sheet_name, index=data_sheet_index)
            print(f"    -> [Service:MasterReport] '{sheet_name}' 시트 생성")
            
            ws.append([])  # A1 빈 행
            ws.append(list(new_data.columns))  # A2 헤더
            for row in dataframe_to_rows(new_data, index=False, header=False):
                ws.append(row)
    
    def _create_pivot_sheet(
        self,
        book: openpyxl.Workbook,
        sheet_name: str,
        pivot_sheet_name: str,
        pivot_data: pd.DataFrame,
        date_int: int
    ) -> None:
        """피벗 시트를 생성하고 서식을 적용합니다."""
        # 기존 피벗 시트 삭제
        if pivot_sheet_name in book.sheetnames:
            book.remove(book[pivot_sheet_name])
        
        # Raw 시트 앞에 피벗 시트 생성
        try:
            data_sheet_index = book.sheetnames.index(sheet_name)
        except ValueError:
            data_sheet_index = 0
        
        pivot_ws = book.create_sheet(title=pivot_sheet_name, index=data_sheet_index)
        print(f"    -> [Service:MasterReport] '{pivot_sheet_name}' 피벗 시트 생성")
        
        # 서식 정의
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        red_font = Font(color="FF0000")
        top_5_colors = ["FF0000", "FFC000", "FFFF00", "92D050", "00B0F0"]
        top_5_fills = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in top_5_colors]
        
        # 2행 비우기
        pivot_ws.append([])
        pivot_ws.append([])
        
        # A열 너비 조정
        pivot_ws.column_dimensions['A'].width = 22.86
        
        # 피벗 데이터 쓰기
        if not pivot_data.empty:
            for r in dataframe_to_rows(pivot_data, index=True, header=True):
                pivot_ws.append(r)
            
            max_col = 1 + len(pivot_data.columns)
            data_start_row = 5
            
            # 헤더 배경색
            for row in pivot_ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=max_col):
                for cell in row:
                    cell.fill = header_fill
            
            # Top 20 빨간색 폰트
            safe_end_row = min(data_start_row + 19, pivot_ws.max_row)
            for row in pivot_ws.iter_rows(min_row=data_start_row, max_row=safe_end_row, min_col=1, max_col=1):
                row[0].font = red_font
            
            # 당일 Top 5 배경색
            if date_int in pivot_data.columns:
                try:
                    # 총계 제외한 피벗에서 날짜 열 찾기
                    pivot_without_total = pivot_data.drop(columns=['총계']) if '총계' in pivot_data.columns else pivot_data
                    if date_int in pivot_without_total.columns:
                        date_col_idx = list(pivot_without_total.columns).index(date_int)
                        target_col = date_col_idx + 2  # 인덱스 열 고려
                        
                        top_5_series = pivot_data[date_int].nlargest(5)
                        top_5_series = top_5_series[top_5_series > 0]
                        top_5_map = {stock: fill for stock, fill in zip(top_5_series.index, top_5_fills)}
                        
                        if top_5_map:
                            print(f"    -> [Service:MasterReport] 당일 Top {len(top_5_map)} 배경색 적용")
                            for row in pivot_ws.iter_rows(min_row=data_start_row, max_row=pivot_ws.max_row, min_col=1, max_col=target_col):
                                if row[0].value in top_5_map:
                                    row[target_col - 1].fill = top_5_map[row[0].value]
                except Exception as e:
                    print(f"    -> [Service:MasterReport] ⚠️ 배경색 적용 건너뜀: {e}")
