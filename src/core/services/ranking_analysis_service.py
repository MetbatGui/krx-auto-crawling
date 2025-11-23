"""
순위표 생성 및 분석 비즈니스 로직 서비스

전체 워크플로우를 오케스트레이션하고 StoragePort를 통해 파일 I/O를 수행합니다.
"""
import pandas as pd
import datetime
from typing import Dict, Set, List, Optional, Any
from pathlib import Path

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import FILL_NONE

from core.ports.storage_port import StoragePort
from core.domain.models import KrxData


class RankingAnalysisService:
    """순위표 생성 및 분석 서비스"""
    
    TOP_N = 20
    COMMON_STOCK_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    LAYOUT_MAP = {
        'KOSPI_foreigner': {'stock_col': 'D', 'value_col': 'E', 'start_row': 5, 'market': 'KOSPI'},
        'KOSPI_institutions': {'stock_col': 'F', 'value_col': 'G', 'start_row': 5, 'market': 'KOSPI'},
        'KOSDAQ_foreigner': {'stock_col': 'I', 'value_col': 'J', 'start_row': 5, 'market': 'KOSDAQ'},
        'KOSDAQ_institutions': {'stock_col': 'K', 'value_col': 'L', 'start_row': 5, 'market': 'KOSDAQ'},
    }
    DATA_RANGE_TO_CLEAR = "D5:L24"
    COLUMNS_TO_AUTOFIT = ['D', 'F', 'I', 'K']
    
    def __init__(self, storage: StoragePort, file_name: str = "2025일별수급순위정리표.xlsx"):
        """
        Args:
            storage: 파일 저장/로드를 위한 StoragePort
            file_name: Excel 파일명
        """
        self.storage = storage
        self.file_path = file_name
        self.korean_weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        print(f"[Service:RankingAnalysis] 초기화 완료 (파일: {self.file_path})")
    
    def update_ranking_report(self, data_list: List[KrxData]) -> None:
        """
        순위표 전체 업데이트 워크플로우
        
        Args:
            data_list: 업데이트할 KRX 데이터 리스트
        """
        if not data_list:
            print("[Service:RankingAnalysis] ⚠️ 데이터가 없어 건너뜁니다")
            return
        
        # 1. 데이터 변환
        data_map = {item.key: item.data for item in data_list if not item.data.empty}
        
        # 2. 날짜 추출
        report_date = datetime.datetime.strptime(data_list[0].date_str, '%Y%m%d').date()
        
        # 3. 공통 종목 계산
        common_stocks = self.calculate_common_stocks(data_map)
        
        # 4. 리포트 업데이트
        self._execute_update(report_date, data_map, common_stocks)
    
    def calculate_common_stocks(self, data_map: Dict[str, pd.DataFrame]) -> Dict[str, Set[str]]:
        """
        시장별 외국인/기관 공통 매수 종목을 계산합니다.
        
        Args:
            data_map: {key: DataFrame} 형태의 데이터
            
        Returns:
            시장별 공통 종목 Set
        """
        common_stocks = {}
        markets = ['KOSPI', 'KOSDAQ']
        
        for market in markets:
            foreigner_key = f"{market}_foreigner"
            inst_key = f"{market}_institutions"
            
            df_foreign = data_map.get(foreigner_key)
            df_inst = data_map.get(inst_key)
            
            if df_foreign is not None and df_inst is not None:
                top_foreign = set(df_foreign.head(self.TOP_N)['종목명'])
                top_inst = set(df_inst.head(self.TOP_N)['종목명'])
                
                common = top_foreign.intersection(top_inst)
                common_stocks[market] = common
                
                print(f"    -> [Service:RankingAnalysis] {market} 공통 종목 ({len(common)}개): {common}")
            else:
                common_stocks[market] = set()
                print(f"    -> [Service:RankingAnalysis] {market} 데이터 부족")
        
        return common_stocks
    
    def _execute_update(
        self,
        report_date: datetime.date,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ) -> bool:
        """전체 업데이트 실행"""
        print(f"    -> [Service:RankingAnalysis] 순위표 업데이트 시작...")
        
        # 1. 워크북 로드
        book = self.storage.load_workbook(self.file_path)
        if book is None:
            print(f"    -> [Service:RankingAnalysis] 🚨 파일을 찾을 수 없습니다")
            return False
        
        # 2. 템플릿 시트 찾기
        if not book.worksheets:
            print(f"    -> [Service:RankingAnalysis] 🚨 시트가 없습니다")
            return False
        source_sheet = book.worksheets[-1]
        
        # 3. 시트 복사 및 준비
        try:
            new_sheet = self._copy_and_prepare_sheet(book, source_sheet, report_date)
        except Exception as e:
            print(f"    -> [Service:RankingAnalysis] 🚨 시트 복사 실패: {e}")
            return False
        
        # 4. 헤더 업데이트
        try:
            self._update_sheet_headers(new_sheet, report_date)
        except Exception as e:
            print(f"    -> [Service:RankingAnalysis] 🚨 헤더 업데이트 실패: {e}")
            return False
        
        # 5. 데이터 붙여넣기 및 서식
        try:
            self._paste_and_format_data(new_sheet, data_map, common_stocks)
        except Exception as e:
            print(f"    -> [Service:RankingAnalysis] 🚨 데이터 적용 실패: {e}")
            return False
        
        # 6. 저장
        success = self.storage.save_workbook(book, self.file_path)
        if success:
            print(f"    -> [Service:RankingAnalysis] ✅ 순위표 저장 완료")
        return success
    
    def _copy_and_prepare_sheet(
        self,
        book: Workbook,
        source_sheet: Worksheet,
        report_date: datetime.date
    ) -> Worksheet:
        """시트 복사 및 준비"""
        sheet_name = report_date.strftime('%m%d')
        
        # 기존 시트 삭제
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        
        # 시트 복사
        new_sheet = book.copy_worksheet(source_sheet)
        new_sheet.title = sheet_name
        
        print(f"    -> [Service:RankingAnalysis] '{sheet_name}' 시트 생성")
        return new_sheet
    
    def _update_sheet_headers(self, sheet: Worksheet, report_date: datetime.date) -> None:
        """헤더 업데이트"""
        sheet['A5'] = report_date.strftime('%Y-%m-%d')
        weekday_idx = report_date.weekday()
        sheet['B5'] = self.korean_weekdays[weekday_idx]
    
    def _paste_and_format_data(
        self,
        ws: Worksheet,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ) -> None:
        """데이터 붙여넣기 및 서식 적용"""
        # 1. 데이터 영역 초기화
        self._clear_data_area(ws)
        
        # 2. 각 리스트별 데이터 붙여넣기
        for key, layout in self.LAYOUT_MAP.items():
            df = data_map.get(key)
            if df is not None and not df.empty:
                pasted_count = self._paste_single_list(ws, df, layout)
                
                # 공통 종목 서식
                market = layout['market']
                if market in common_stocks:
                    self._apply_common_stock_format(
                        ws, layout, common_stocks[market], pasted_count
                    )
                
                # 남은 행 정리
                self._clear_remaining_rows(ws, layout, pasted_count)
        
        # 3. 자동 너비 맞춤
        self._apply_autofit(ws)
    
    def _clear_data_area(self, ws: Worksheet) -> None:
        """데이터 영역 초기화"""
        for row in ws[self.DATA_RANGE_TO_CLEAR]:
            for cell in row:
                cell.value = None
                cell.fill = FILL_NONE
    
    def _paste_single_list(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        layout: Dict[str, Any]
    ) -> int:
        """단일 리스트 데이터 붙여넣기"""
        stock_col = layout['stock_col']
        value_col = layout['value_col']
        start_row = layout['start_row']
        
        count = min(len(df), self.TOP_N)
        for i in range(count):
            row_num = start_row + i
            ws[f"{stock_col}{row_num}"] = df.iloc[i]['종목명']
            ws[f"{value_col}{row_num}"] = df.iloc[i]['순매수_거래대금']
        
        return count
    
    def _apply_common_stock_format(
        self,
        ws: Worksheet,
        layout: Dict[str, Any],
        common_set: Set[str],
        count: int
    ) -> None:
        """공통 종목 서식 적용"""
        stock_col = layout['stock_col']
        start_row = layout['start_row']
        
        for i in range(count):
            row_num = start_row + i
            stock_name = ws[f"{stock_col}{row_num}"].value
            if stock_name in common_set:
                ws[f"{stock_col}{row_num}"].fill = self.COMMON_STOCK_FILL
    
    def _clear_remaining_rows(
        self,
        ws: Worksheet,
        layout: Dict[str, Any],
        count: int
    ) -> None:
        """남은 행 정리"""
        stock_col = layout['stock_col']
        value_col = layout['value_col']
        start_row = layout['start_row']
        
        for i in range(count, self.TOP_N):
            row_num = start_row + i
            ws[f"{stock_col}{row_num}"].value = None
            ws[f"{value_col}{row_num}"].value = None
    
    def _apply_autofit(self, ws: Worksheet) -> None:
        """자동 너비 맞춤"""
        for col in self.COLUMNS_TO_AUTOFIT:
            ws.column_dimensions[col].bestFit = True
            ws.column_dimensions[col].auto_size = True
