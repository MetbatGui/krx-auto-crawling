"""
Excel 서식 적용 유틸리티

MasterExcelAdapter와 RankingExcelAdapter에서 공통으로 사용하는 
서식 관련 로직을 제공합니다.
"""
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Set


class ExcelFormatter:
    """Excel 서식 적용 유틸리티 클래스"""
    
    # 공통 색상 정의
    COLORS = {
        'header_blue': 'DDEBF7',      # 헤더 파란색
        'common_blue': 'B4C6E7',      # 공통 종목 파란색
        'red': 'FF0000',              # 빨강
        'orange': 'FFC000',           # 주황
        'yellow': 'FFFF00',           # 노랑
        'green': '92D050',            # 초록
        'light_blue': '00B0F0',       # 하늘색
    }
    
    @staticmethod
    def apply_header_fill(
        ws: Worksheet,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
        color: str = 'header_blue'
    ):
        """헤더 영역에 배경색을 적용합니다.
        
        Args:
            ws (Worksheet): 워크시트.
            min_row (int): 시작 행.
            max_row (int): 종료 행.
            min_col (int): 시작 열.
            max_col (int): 종료 열.
            color (str): 색상 키 (기본: 'header_blue').
        """
        fill = PatternFill(
            start_color=ExcelFormatter.COLORS[color],
            end_color=ExcelFormatter.COLORS[color],
            fill_type="solid"
        )
        
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col
        ):
            for cell in row:
                cell.fill = fill
    
    @staticmethod
    def apply_font_color(
        ws: Worksheet,
        min_row: int,
        max_row: int,
        col: int,
        color: str = 'red'
    ):
        """특정 열의 폰트 색상을 변경합니다.
        
        Args:
            ws (Worksheet): 워크시트.
            min_row (int): 시작 행.
            max_row (int): 종료 행.
            col (int): 열 번호.
            color (str): 색상 키 (기본: 'red').
        """
        font = Font(color=ExcelFormatter.COLORS[color])
        
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=col,
            max_col=col
        ):
            row[0].font = font
    
    @staticmethod
    def apply_top_backgrounds(
        ws: Worksheet,
        start_row: int,
        date_col: str,
        top_5_stocks: List[str]
    ):
        """상위 5개 종목에 그라데이션 배경색을 적용합니다.
        
        빨강 → 주황 → 노랑 → 초록 → 하늘색 순서로 적용됩니다.
        
        Args:
            ws (Worksheet): 워크시트.
            start_row (int): 데이터 시작 행.
            date_col (str): 날짜 열 (예: 'B', 'C').
            top_5_stocks (List[str]): 상위 5개 종목명 리스트.
        """
        top_5_colors = ['red', 'orange', 'yellow', 'green', 'light_blue']
        
        for i, stock_name in enumerate(top_5_stocks[:5]):
            if i >= 5:
                break
            
            color_key = top_5_colors[i]
            fill = PatternFill(
                start_color=ExcelFormatter.COLORS[color_key],
                end_color=ExcelFormatter.COLORS[color_key],
                fill_type="solid"
            )
            
            # 해당 종목이 있는 행 찾기
            for row in ws.iter_rows(min_row=start_row, min_col=1, max_col=1):
                cell = row[0]
                if cell.value == stock_name:
                    # 해당 날짜 열의 셀에 배경색 적용
                    ws[f"{date_col}{cell.row}"].fill = fill
                    break
    
    @staticmethod
    def apply_common_stock_fill(
        ws: Worksheet,
        stock_col: str,
        start_row: int,
        pasted_count: int,
        common_stocks: Set[str],
        color: str = 'common_blue'
    ):
        """공통 종목에 배경색을 적용합니다 (RankingExcelAdapter용).
        
        Args:
            ws (Worksheet): 워크시트.
            stock_col (str): 종목명 열 (예: 'D', 'F').
            start_row (int): 시작 행.
            pasted_count (int): 붙여넣은 행 수.
            common_stocks (Set[str]): 공통 종목명 Set.
            color (str): 색상 키 (기본: 'common_blue').
        """
        fill = PatternFill(
            start_color=ExcelFormatter.COLORS[color],
            end_color=ExcelFormatter.COLORS[color],
            fill_type="solid"
        )
        
        for i in range(pasted_count):
            current_row = start_row + i
            stock_cell = ws[f"{stock_col}{current_row}"]
            if stock_cell.value in common_stocks:
                stock_cell.fill = fill
    
    @staticmethod
    def set_column_width(ws: Worksheet, column_letter: str, width: float):
        """열 너비를 설정합니다.
        
        Args:
            ws (Worksheet): 워크시트.
            column_letter (str): 열 문자 (예: 'A', 'B').
            width (float): 너비 값.
        """
        ws.column_dimensions[column_letter].width = width
