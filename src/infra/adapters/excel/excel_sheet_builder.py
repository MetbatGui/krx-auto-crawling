"""
Excel 시트 구성 유틸리티

워크북 내 시트 생성, 데이터 삽입, 구조 구성 등의 로직을 제공합니다.
"""
import pandas as pd
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Optional
import datetime


class ExcelSheetBuilder:
    """Excel 시트 구성 유틸리티 클래스"""
    
    @staticmethod
    def build_data_sheet(
        book: Workbook,
        sheet_name: str,
        data_df: pd.DataFrame,
        sheet_exists: bool,
        index: Optional[int] = None
    ) -> Worksheet:
        """
        데이터 시트를 생성하거나 업데이트합니다.
        
        Args:
            book: Workbook 객체
            sheet_name: 시트명
            data_df: 삽입할 DataFrame
            sheet_exists: 시트가 이미 존재하는지 여부
            index: 시트를 삽입할 위치 (None이면 끝에 추가)
            
        Returns:
            생성되거나 업데이트된 Worksheet
        """
        if sheet_exists:
            ws = book[sheet_name]
        else:
            if index is not None:
                ws = book.create_sheet(title=sheet_name, index=index)
            else:
                ws = book.create_sheet(title=sheet_name)
        
        # DataFrame 데이터를 시트에 작성
        for row in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(row)
        
        return ws
    
    @staticmethod
    def build_pivot_sheet(
        book: Workbook,
        sheet_name: str,
        pivot_df: pd.DataFrame,
        index: Optional[int] = None
    ) -> Worksheet:
        """
        피벗 시트를 생성합니다 (2행 띄우고 시작).
        
        Args:
            book: Workbook 객체
            sheet_name: 시트명
            pivot_df: 피벗 DataFrame
            index: 시트를 삽입할 위치 (None이면 끝에 추가)
            
        Returns:
            생성된 Worksheet
        """
        # 기존 시트가 있으면 제거
        if sheet_name in book.sheetnames:
            book.remove(book[sheet_name])
        
        # 새 시트 생성
        if index is not None:
            ws = book.create_sheet(title=sheet_name, index=index)
        else:
            ws = book.create_sheet(title=sheet_name)
        
        # 2행 띄우기 (A1, A2 비워둠)
        ws.append([])  # 1행
        ws.append([])  # 2행
        
        # 피벗 데이터 작성 (A3부터)
        for row in dataframe_to_rows(pivot_df, index=True, header=True):
            ws.append(row)
        
        return ws
    
    @staticmethod
    def build_ranking_sheet(
        book: Workbook,
        source_sheet: Worksheet,
        report_date: datetime.date,
        data_map: dict,
        layout_map: dict,
        top_n: int = 20
    ) -> Worksheet:
        """
        순위표 시트를 생성합니다 (템플릿 복사 후 데이터 삽입).
        
        Args:
            book: Workbook 객체
            source_sheet: 템플릿 시트
            report_date: 리포트 날짜
            data_map: 데이터 딕셔너리 {key: DataFrame}
            layout_map: 레이아웃 정보 딕셔너리
            top_n: 상위 몇 개 종목을 붙여넣을지
            
        Returns:
            생성된 Worksheet
        """
        # 날짜로 시트명 생성
        new_sheet_name = report_date.strftime('%m%d')
        
        # 동일 시트명이 있으면 제거
        if new_sheet_name in book.sheetnames:
            book.remove(book[new_sheet_name])
        
        # 시트 복사
        new_sheet = book.copy_worksheet(source_sheet)
        new_sheet.title = new_sheet_name
        
        return new_sheet
    
    @staticmethod
    def paste_ranking_data(
        ws: Worksheet,
        df: pd.DataFrame,
        layout: dict,
        top_n: int = 20
    ) -> int:
        """
        순위표 데이터를 시트에 붙여넣습니다.
        
        Args:
            ws: Worksheet
            df: 데이터 DataFrame
            layout: 레이아웃 정보 {'stock_col': 'D', 'value_col': 'E', 'start_row': 5}
            top_n: 상위 몇 개
            
        Returns:
            실제 붙여넣은 행 수
        """
        stock_col = layout['stock_col']
        value_col = layout['value_col']
        start_row = layout['start_row']
        
        # 상위 N개 가져오기
        df_top_n = df.head(top_n)
        
        row_index = 0
        for _, row_data in df_top_n.iterrows():
            current_row = start_row + row_index
            ws[f"{stock_col}{current_row}"].value = row_data['종목명']
            ws[f"{value_col}{current_row}"].value = row_data['순매수_거래대금']
            row_index += 1
        
        return row_index
    
    @staticmethod
    def clear_ranking_remaining_rows(
        ws: Worksheet,
        layout: dict,
        pasted_count: int,
        total_rows: int = 20
    ):
        """
        순위표에서 데이터가 없는 남은 행을 지웁니다.
        
        Args:
            ws: Worksheet
            layout: 레이아웃 정보
            pasted_count: 실제 붙여넣은 행 수
            total_rows: 전체 행 수 (기본 20)
        """
        stock_col = layout['stock_col']
        value_col = layout['value_col']
        start_row = layout['start_row']
        
        for i in range(pasted_count, total_rows):
            current_row = start_row + i
            ws[f"{stock_col}{current_row}"].value = None
            ws[f"{value_col}{current_row}"].value = None
