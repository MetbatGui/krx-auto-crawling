"""
마스터 리포트 피벗 시트 어댑터

openpyxl을 사용하여 피벗 시트 생성 및 서식 적용
"""
import pandas as pd
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill


class MasterPivotSheetAdapter:
    """피벗 시트 생성 및 서식 적용 전용 어댑터"""
    
    def create_pivot_sheet(
        self,
        book: Workbook,
        sheet_name: str,
        pivot_sheet_name: str,
        pivot_data: pd.DataFrame,
        date_int: int
    ) -> None:
        """피벗 시트를 생성하고 서식을 적용합니다.
        
        Args:
            book (Workbook): openpyxl Workbook.
            sheet_name (str): Raw 데이터 시트 이름.
            pivot_sheet_name (str): 피벗 시트 이름.
            pivot_data (pd.DataFrame): 피벗 데이터.
            date_int (int): 기준 날짜.
        """
        # 기존 피벗 시트 삭제
        if pivot_sheet_name in book.sheetnames:
            book.remove(book[pivot_sheet_name])
        
        # Raw 시트 앞에 피벗 시트 생성
        try:
            data_sheet_index = book.sheetnames.index(sheet_name)
        except ValueError:
            data_sheet_index = 0
        
        pivot_ws = book.create_sheet(title=pivot_sheet_name, index=data_sheet_index)
        print(f"    -> [Adapter:MasterPivotSheet] '{pivot_sheet_name}' 피벗 시트 생성")
        
        # 서식 정의
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        red_font = Font(color="FF0000")
        top_5_colors = ["FF0000", "FFC000", "FFFF00", "92D050", "00B0F0"]
        top_5_fills = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in top_5_colors]
        
        # 2행 비우기
        pivot_ws.append([])
        pivot_ws.append([])
        
        # A열 너비 조정
        pivot_ws.column_dimensions['A'].width = 22.86
        
        # 피벗 데이터 쓰기
        if not pivot_data.empty:
            for r in dataframe_to_rows(pivot_data, index=True, header=True):
                pivot_ws.append(r)
            
            max_col = 1 + len(pivot_data.columns)
            data_start_row = 5
            
            # 서식 적용
            self._apply_header_format(pivot_ws, max_col, header_fill)
            self._apply_top20_format(pivot_ws, data_start_row, red_font)
            self._apply_top5_format(pivot_ws, pivot_data, date_int, data_start_row, top_5_fills)
            
            print(f"    -> [Adapter:MasterPivotSheet] 피벗 시트 서식 적용 완료")
        else:
            print(f"    -> [Adapter:MasterPivotSheet] ⚠️ 빈 피벗 시트 생성")
    
    def _apply_header_format(self, ws, max_col, header_fill):
        """헤더 배경색 및 날짜 포맷 적용"""
        for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = header_fill
                # 4행의 날짜 컬럼들에 대해 포맷 적용 (B열부터, 총계 제외)
                # 단순하게 모든 4행 헤더에 대해 적용 시도 (텍스트여도 무방함)
                if row[0].row == 4:
                    cell.number_format = 'yyyymmdd'
    
    def _apply_top20_format(self, ws, data_start_row, red_font):
        """Top 30 빨간색 폰트 적용"""
        safe_end_row = min(data_start_row + 29, ws.max_row)
        for row in ws.iter_rows(min_row=data_start_row, max_row=safe_end_row, min_col=1, max_col=1):
            row[0].font = red_font
    
    def _apply_top5_format(self, ws, pivot_data, date_int, data_start_row, top_5_fills):
        """당일 Top 5 배경색 적용"""
        if date_int not in pivot_data.columns:
            return
        
        try:
            # 총계 제외한 피벗에서 날짜 열 찾기
            pivot_without_total = pivot_data.drop(columns=['총계']) if '총계' in pivot_data.columns else pivot_data
            if date_int not in pivot_without_total.columns:
                return
            
            date_col_idx = list(pivot_without_total.columns).index(date_int)
            target_col = date_col_idx + 2  # 인덱스 열 고려
            
            top_5_series = pivot_data[date_int].nlargest(5)
            top_5_series = top_5_series[top_5_series > 0]
            top_5_map = {stock: fill for stock, fill in zip(top_5_series.index, top_5_fills)}
            
            if top_5_map:
                print(f"    -> [Adapter:MasterPivotSheet] 당일 Top {len(top_5_map)} 배경색 적용")
                for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, min_col=1, max_col=target_col):
                    if row[0].value in top_5_map:
                        row[target_col - 1].fill = top_5_map[row[0].value]
        except Exception as e:
            print(f"    -> [Adapter:MasterPivotSheet] ⚠️ 배경색 적용 건너뜀: {e}")
