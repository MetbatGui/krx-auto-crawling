"""
마스터 리포트 Raw 데이터 시트 어댑터

openpyxl을 사용하여 Raw 데이터 시트 생성 및 업데이트
"""
import pandas as pd
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class MasterSheetAdapter:
    """Raw 데이터 시트 생성 및 업데이트 전용 어댑터"""
    
    def update_raw_sheet(
        self,
        book: Workbook,
        sheet_name: str,
        new_data: pd.DataFrame,
        sheet_exists: bool
    ) -> None:
        """Raw 데이터 시트를 업데이트합니다.
        
        Args:
            book (Workbook): openpyxl Workbook.
            sheet_name (str): 시트 이름.
            new_data (pd.DataFrame): 추가할 데이터.
            sheet_exists (bool): 시트 존재 여부.
        """
        if sheet_exists and sheet_name in book.sheetnames:
            # 기존 시트에 추가
            ws = book[sheet_name]
            print(f"    -> [Adapter:MasterSheet] '{sheet_name}' 시트에 데이터 추가")
            for row in dataframe_to_rows(new_data, index=False, header=False):
                ws.append(row)
        else:
            # 새 시트 생성 (마지막 시트 앞에)
            data_sheet_index = max(0, len(book.sheetnames) - 1) if book.sheetnames else 0
            ws = book.create_sheet(title=sheet_name, index=data_sheet_index)
            print(f"    -> [Adapter:MasterSheet] '{sheet_name}' 시트 생성")
            
            ws.append([])  # A1 빈 행
            ws.append(list(new_data.columns))  # A2 헤더
            for row in dataframe_to_rows(new_data, index=False, header=False):
                ws.append(row)
        
        print(f"    -> [Adapter:MasterSheet] Raw 시트 업데이트 완료")
