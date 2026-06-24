"""
마스터 리포트 워크북 어댑터

SheetAdapter와 PivotSheetAdapter를 조합하여 완전한 워크북 생성
"""
import pandas as pd
import openpyxl
from openpyxl.workbook.workbook import Workbook
from typing import Optional, List

from core.ports.storage_port import StoragePort
from infra.adapters.excel.master_sheet_adapter import MasterSheetAdapter
from infra.adapters.excel.master_pivot_sheet_adapter import MasterPivotSheetAdapter


class MasterWorkbookAdapter:
    """마스터 리포트 워크북 조립 및 저장 전용 어댑터"""
    
    def __init__(
        self,
        source_storage: StoragePort,
        target_storages: List[StoragePort],
        sheet_adapter: MasterSheetAdapter,
        pivot_sheet_adapter: MasterPivotSheetAdapter
    ):
        self.source_storage = source_storage
        self.target_storages = target_storages
        self.sheet_adapter = sheet_adapter
        self.pivot_sheet_adapter = pivot_sheet_adapter
    
    def save_workbook(
        self,
        file_path: str,
        sheet_name: str,
        pivot_sheet_name: str,
        new_data: pd.DataFrame,
        pivot_data: pd.DataFrame,
        date_int: int,
        sheet_exists: bool
    ) -> bool:
        """워크북을 생성하고 저장합니다.
        
        Args:
            file_path (str): 파일 경로 (상대 경로).
            sheet_name (str): Raw 데이터 시트 이름.
            pivot_sheet_name (str): 피벗 시트 이름.
            new_data (pd.DataFrame): 추가할 새 데이터.
            pivot_data (pd.DataFrame): 피벗 테이블 데이터.
            date_int (int): 기준 날짜.
            sheet_exists (bool): Raw 시트 존재 여부.
            
        Returns:
            bool: 저장 성공 여부 (모두 성공 시 True).
        """
        try:
            # 1. 워크북 로드 또는 생성 (Source Storage 사용)
            book = self.source_storage.load_workbook(file_path)
            if book is None:
                book = openpyxl.Workbook()
                if 'Sheet' in book.sheetnames:
                    book.remove(book['Sheet'])
            
            # 2. Raw 데이터 시트 업데이트
            if not new_data.empty:
                self.sheet_adapter.update_raw_sheet(book, sheet_name, new_data, sheet_exists)
            
            # 3. 피벗 시트 생성
            self.pivot_sheet_adapter.create_pivot_sheet(
                book, sheet_name, pivot_sheet_name,
                pivot_data, date_int
            )
            
            # 4. 저장 (Target Storages 모두에 저장)
            all_success = True
            for storage in self.target_storages:
                success = storage.save_workbook(book, file_path)
                if success:
                    print(f"    -> [Adapter:MasterWorkbook] [OK] {storage.__class__.__name__} 저장 완료")
                    if not pivot_data.empty:
                        # 로그는 한 번만 출력하거나 저장소별로 출력
                        pass
                else:
                    all_success = False
            
            if not pivot_data.empty:
                 print(f"    -> [Adapter:MasterWorkbook] 피벗 샘플:\n{pivot_data.head()}")

            return all_success
            
        except Exception as e:
            print(f"    -> [Adapter:MasterWorkbook] [Error] 워크북 저장 실패: {e}")
            return False
