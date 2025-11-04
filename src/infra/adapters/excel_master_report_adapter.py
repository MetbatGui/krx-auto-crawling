import pandas as pd
import datetime
from typing import Dict
import os

# (pip install openpyxl)
import openpyxl
# [V9/V13] dataframe_to_rows 임포트
from openpyxl.utils.dataframe import dataframe_to_rows
# [V16-V21] 서식 적용을 위한 Font, PatternFill 임포트
from openpyxl.styles import Font, PatternFill

from core.ports.excel_master_report_port import ExcelMasterReportPort

class ExcelMasterAdapter(ExcelMasterReportPort):
    """
    ExcelMasterReportPort의 구현체(Adapter).
    
    [최종 로직 V22 - 빠른 건너뛰기]
    1. (V22) 모든 작업 시작 전, 파일 존재 여부와 
       'pivot_sheet_name' 존재 여부를 "먼저" 확인하고,
       시트가 이미 있으면 [빠른 건너뛰기]로 즉시 True를 반환.
    2. (V15) I/O 최적화 로직 유지
    3. (V19) 시트 순서 보장
    4. (V20-V21) 모든 서식 적용
    """

    def __init__(self, base_path: str, file_name_prefix: str = "2025"):
        # ('output/순매수도' 폴더 생성)
        self.master_path = os.path.join(base_path, "순매수도")
        if not os.path.exists(self.master_path):
            os.makedirs(self.master_path)
            
        # (파일명 형식: '...순매수도(2025).xlsx')
        year_suffix = f"({file_name_prefix})"
        self.file_map: Dict[str, str] = {
            'KOSPI_foreigner': f'코스피외국인순매수도{year_suffix}.xlsx',
            'KOSDAQ_foreigner': f'코스닥외국인순매수도{year_suffix}.xlsx',
            'KOSPI_institutions': f'코스피기관순매수도{year_suffix}.xlsx',
            'KOSDAQ_institutions': f'코스닥기관순매수도{year_suffix}.xlsx',
        }

    def update_report(
        self,
        report_key: str,
        daily_data: pd.DataFrame,
        report_date: datetime.date
    ) -> bool:
        
        file_name = self.file_map.get(report_key)
        if not file_name:
            print(f"    -> [Adapter] 🚨 '{report_key}'에 해당하는 파일명을 모릅니다.")
            return False

        file_path = os.path.join(self.master_path, file_name)
        
        # [V13] 시트 이름 정의
        sheet_name = report_date.strftime('%b').upper()
        pivot_sheet_name = report_date.strftime('%m%d') 
        
        date_str = report_date.strftime('%Y%m%d')
        date_int = int(date_str) 

        print(f"    -> [Adapter] {file_name} 파일 업데이트 시작...")
        print(f"         (1단계: '{sheet_name}' 누적, 2단계: '{pivot_sheet_name}' 피벗 생성)")

        # --- [V22] 빠른 건너뛰기 로직 ---
        # 1. 파일을 열기 전에, 파일이 존재하는지 확인
        if os.path.exists(file_path):
            try:
                # 2. 존재한다면, 시트 이름만 빠르게 읽어옴 (read_only=True)
                #    (파일이 깨졌을 경우를 대비해 try/except)
                book = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = book.sheetnames
                book.close()
                
                # 3. 오늘 날짜의 피벗 시트가 이미 있는지 확인
                if pivot_sheet_name in sheet_names:
                    print(f"    -> [Adapter] ⚠️ '{pivot_sheet_name}' 피벗 시트가 이미 존재하여 [빠른 건너뛰기]를 실행합니다.")
                    return True # True를 반환하여 파이프라인 계속 진행
            except Exception as e:
                # (예: 파일이 깨졌거나, zip 파일이 아닌 경우)
                print(f"    -> [Adapter] ⚠️ 빠른 건너뛰기 검사 중 파일({file_name})을 읽을 수 없습니다: {e}")
                print(f"    -> [Adapter] ⚠️ (파일을 덮어쓰기 위해, 전체 로직을 계속 진행합니다.)")
                pass # 에러가 났으므로, 정상 로직을 태워서 덮어쓰도록 유도
        # --- [V22] 빠른 건너뛰기 끝 ---


        # --- (V22) (여기까지 왔다면, 피벗 시트가 없거나 파일이 없으므로, V21의 전체 로직을 실행) ---


        # --- 1. [V11] 새 데이터를 엑셀 스키마로 번역 ---
        try:
            data_dict = {
                '일자': date_int, 
                '종목': daily_data['종목명'],
                '금액': pd.to_numeric(daily_data['순매수_거래대금'])
            }
            new_data_formatted = pd.DataFrame(data_dict)
            new_data_formatted = new_data_formatted[['일자', '종목', '금액']]

        except KeyError as e:
            print(f"    -> [Adapter] 🚨 'daily_data'에 필요한 컬럼이 없습니다: {e}")
            return False

        # --- 2. [V15] 기존 데이터 읽기 (Pandas, 중복 검사용 - 1회 읽기) ---
        excel_columns = ['일자', '종목', '금액']
        existing_df = pd.DataFrame(columns=excel_columns)
        sheet_exists = False 
        try:
            read_df = pd.read_excel(
                file_path, 
                sheet_name=sheet_name, 
                engine='openpyxl', 
                skiprows=1,
                dtype={'일자': int}
            )
            sheet_exists = True 
            if not read_df.empty:
                if all(col in read_df.columns for col in read_df.columns):
                        existing_df = read_df[excel_columns].copy()
                else:
                    print(f"    -> [Adapter] ⚠️ {sheet_name} 시트 헤더가 깨져 읽을 수 없습니다.")
                    existing_df = pd.DataFrame(columns=excel_columns)
            print(f"    -> [Adapter] 기존 '{sheet_name}' 시트 데이터 ({len(existing_df)}줄) 로드 완료.")
        except FileNotFoundError:
            print(f"    -> [Adapter] ⚠️ 새 파일 '{file_name}'이 생성됩니다.")
        except (ValueError, KeyError) as e:
            print(f"    -> [Adapter] ⚠️ 파일은 있으나 '{sheet_name}' 시트가 없어 새로 생성합니다.")
        except Exception as e:
            print(f"    -> [Adapter] 🚨 파일 로드 중 예상치 못한 오류: {e}")
            return False

        # --- 3. [V13 수정] 중복 날짜 검사 ---
        if date_int in existing_df['일자'].values: 
            print(f"    -> [Adapter] ⚠️ {date_int} 데이터가 '{sheet_name}'에 이미 존재하여 무시합니다.")
            new_data_formatted = pd.DataFrame()
            print("         (데이터 추가는 건너뛰고, 피벗 테이블 생성(2단계)은 진행합니다.)")
        
        if not new_data_formatted.empty:
            print(f"    -> [Adapter] 새 데이터 ({len(new_data_formatted)}줄) 추가 준비...")

        # --- 4. [V15] 피벗 생성을 위해 메모리에서 전체 데이터 준비 ---
        print(f"    -> [Adapter] 메모리에서 피벗용 전체 데이터 준비...")
        if not new_data_formatted.empty:
            full_data_df = pd.concat([existing_df, new_data_formatted], ignore_index=True)
        else:
            full_data_df = existing_df.copy()

        # --- 5. [V15] 피벗 테이블 계산 (파일 쓰기 전) ---
        print(f"    -> [Adapter] '{pivot_sheet_name}' 피벗 테이블 계산 시작...")
        pivot_df_sorted = pd.DataFrame()
        # 피벗을 만들기 전, '금액' 컬럼을 숫자로 강제 변환합니다.
        # (기존 데이터가 "1,234,000" 처럼 문자열로 로드되었을 경우 대비)
        if not full_data_df.empty:
            try:
                # 1. 쉼표 등 불필요한 문자 제거 (숫자, 소수점, 마이너스 부호 외)
                full_data_df['금액'] = full_data_df['금액'].astype(str).str.replace(r'[^0-9.-]', '', regex=True)
                # 2. 빈 문자열은 0으로
                full_data_df['금액'] = full_data_df['금액'].replace('', 0)
                # 3. 숫자로 변환 (오류 시 NaN)
                full_data_df['금액'] = pd.to_numeric(full_data_df['금액'], errors='coerce')
                # 4. NaN을 0으로 (결측치 방지)
                full_data_df['금액'] = full_data_df['금액'].fillna(0)
            except Exception as clean_e:
                print(f"    -> [Adapter] 🚨 '금액' 컬럼 숫자 변환 중 오류: {clean_e}")
                # (오류가 나도 일단 진행 시도)
        # --- [수정 코드 끝] ---
        if full_data_df.empty:
             print(f"    -> [Adapter] ⚠️ '{sheet_name}' 원본 데이터가 비어있어 피벗을 생성할 수 없습니다.")
        else:
            try:
                # [V21] '총계' 추가 전의 원본 피벗 (오늘 날짜 열 찾기용)
                pivot_df = pd.pivot_table(
                    full_data_df,
                    values='금액',
                    index='종목',
                    columns='일자',
                    aggfunc='sum'
                )
                pivot_df['총계'] = pivot_df.sum(axis=1)
                pivot_df_sorted = pivot_df.sort_values(by='총계', ascending=False)
                print(f"    -> [Adapter] 피벗 테이블 계산 완료.")
            except Exception as e:
                print(f"    -> [Adapter] 🚨 피벗 테이블 계산 중 예외 발생: {e}")
                return False 

        # --- 6. [V22] 엑셀 파일 한 번에 쓰기 (모든 서식/순서 적용) ---
        print(f"    -> [Adapter] 엑셀 파일 쓰기 작업 시작 ({file_name})...")
        try:
            try:
                book = openpyxl.load_workbook(file_path)
            except FileNotFoundError:
                book = openpyxl.Workbook()
                if 'Sheet' in book.sheetnames:
                    book.remove(book['Sheet'])
            
            # (V22) "빠른 건너뛰기"를 통과했으므로, 여기서는 V21-Skip의 중복 검사 로직이 필요 없음

            # --- [쓰기 1단계: 'OCT' 시트 누적 (V19 순서 보장)] ---
            if not new_data_formatted.empty:
                if sheet_exists: 
                    ws = book[sheet_name]
                    print(f"        -> [1단계] '{sheet_name}' 시트 마지막 행({ws.max_row})에 추가합니다.")
                    for row in dataframe_to_rows(new_data_formatted, index=False, header=False):
                        ws.append(row)
                else:
                    # [V19] 'OCT' 시트가 없으면, 마지막 시트(총결산) "앞에" 생성
                    data_sheet_index = 0
                    if len(book.sheetnames) > 0:
                        data_sheet_index = len(book.sheetnames) - 1 
                    
                    ws = book.create_sheet(title=sheet_name, index=data_sheet_index)
                    print(f"        -> [1단계] 새 '{sheet_name}' 시트를 {data_sheet_index}번째 (총결산 앞)에 생성.")
                    ws.append([]) # A1
                    ws.append(list(new_data_formatted.columns)) # A2
                    for row in dataframe_to_rows(new_data_formatted, index=False, header=False):
                        ws.append(row)
                print(f"        -> [1단계] ✅ '{sheet_name}' 시트 누적 완료 (아직 저장 전).")
            else:
                print(f"        -> [1단계] ⏭️ 데이터가 (중복 등으로) 비어있어 누적을 건너뜁니다.")

            # --- [쓰기 2단계: '1023' 피벗 덮어쓰기 및 서식 적용 (V22)] ---
            
            # (V22) V21-Skip과 달리, 피벗 시트가 *혹시라도* 존재하면 (예: 깨진 파일 복구 시)
            # 덮어쓰기 위해 "삭제" 로직을 다시 복원
            if pivot_sheet_name in book.sheetnames:
                print(f"        -> [2단계] ⚠️ (경고) '{pivot_sheet_name}' 시트가 존재하여 덮어씁니다.")
                book.remove(book[pivot_sheet_name])
            
            # [V19] 'OCT' 시트 "바로 앞"에 피벗 시트 생성
            try:
                data_sheet_index = book.sheetnames.index(sheet_name)
            except ValueError:
                if not new_data_formatted.empty: 
                    data_sheet_index = book.sheetnames.index(sheet_name)
                else: 
                    print(f"        -> [2단계] 🚨 버그: '{sheet_name}' 시트를 찾을 수 없습니다. 피벗을 맨 뒤에 생성합니다.")
                    data_sheet_index = -1 
                
            pivot_ws = book.create_sheet(title=pivot_sheet_name, index=data_sheet_index)
            print(f"        -> [2단계] '{pivot_sheet_name}' 시트를 {data_sheet_index}번째 ('{sheet_name}' 앞)에 생성.")

            
            # [V21] 1. 서식 스타일 정의
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # 하늘색
            red_font = Font(color="FF0000") # 빨간색
            top_5_colors = ["FF0000", "FFC000", "FFFF00", "92D050", "00B0F0"]
            top_5_fills = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in top_5_colors]
            
            # [V20] 2. 2행 아래로 내리기 (A1, A2는 비움)
            pivot_ws.append([]) # 빈 행 (1행)
            pivot_ws.append([]) # 빈 행 (2행)
            
            # [V20] 3. A열 너비 조정 (165px -> 22.86)
            pivot_ws.column_dimensions['A'].width = 22.86
            
            # [V20] 4. 피벗 데이터 쓰기 (A3부터) 및 최대 열 계산
            max_col_to_style = 1 
            data_start_row = 5 
            
            if not pivot_df_sorted.empty:
                print(f"        -> [2단계] '{pivot_sheet_name}' 시트 A3셀부터 피벗 데이터 쓰기...")
                max_col_to_style = 1 + len(pivot_df_sorted.columns) 
                
                for r in dataframe_to_rows(pivot_df_sorted, index=True, header=True):
                    pivot_ws.append(r) # A3부터 써짐
                print(f"        -> [2단계] ✅ '{pivot_sheet_name}' 시트 덮어쓰기 완료 (아직 저장 전).")
            else:
                 print(f"        -> [2단계] ⚠️ 피벗 데이터가 비어있어 '{pivot_sheet_name}'는 빈 시트로 생성됩니다.")

            # [V20] 5. A3:?4 (동적) 범위에 배경색 적용 (3행, 4행)
            print(f"        -> [2단계] 헤더 영역 (3-4행, {max_col_to_style}개 열) 배경색 적용...")
            for row in pivot_ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=max_col_to_style):
                for cell in row:
                    cell.fill = header_fill

            # [V21] 6. Top 20 '총계' 종목 (A열) 폰트 빨간색 적용
            if not pivot_df_sorted.empty:
                top_20_end_row = (data_start_row - 1) + 20
                print(f"        -> [2단계] Top 20 (A{data_start_row}:A{top_20_end_row}) 종목 폰트 적용...")
                safe_end_row = min(top_20_end_row, pivot_ws.max_row) 
                for row in pivot_ws.iter_rows(min_row=data_start_row, max_row=safe_end_row, min_col=1, max_col=1):
                    row[0].font = red_font
            
            # [V21] 7. 당일 Top 5 배경색 적용
            if not pivot_df_sorted.empty and (date_int in pivot_df.columns):
                try:
                    date_col_idx_in_df = list(pivot_df.columns).index(date_int)
                    target_col_in_ws = date_col_idx_in_df + 2 
                    top_5_series = pivot_df_sorted[date_int].nlargest(5)
                    top_5_series = top_5_series[top_5_series > 0] 
                    top_5_stock_map = {stock_name: fill for stock_name, fill in zip(top_5_series.index, top_5_fills)}

                    if top_5_stock_map:
                        print(f"        -> [2단계] 당일 Top {len(top_5_stock_map)} ({date_int}, {target_col_in_ws}열) 배경색 적용...")
                        for row in pivot_ws.iter_rows(min_row=data_start_row, max_row=pivot_ws.max_row, min_col=1, max_col=target_col_in_ws):
                            stock_name_cell = row[0] 
                            if stock_name_cell.value in top_5_stock_map:
                                target_cell = row[target_col_in_ws - 1] 
                                target_cell.fill = top_5_stock_map[stock_name_cell.value]
                    else:
                        print(f"        -> [2단계] ⚠️ 당일 Top 5 (양수) 데이터가 없어 배경색을 건너뜁니다.")
                except ValueError:
                     print(f"        -> [2단계] ⚠️ 오늘 날짜({date_int})를 피벗 컬럼에서 찾을 수 없어 배경색을 건너뜁니다.")
            else:
                 print(f"        -> [2단계] ⚠️ 피벗이 비어있거나 오늘 날짜({date_int})가 없어 배경색을 건너뜁니다.")
                 
            # [최종 저장] - 모든 변경사항을 한 번에 저장
            book.save(file_path)
            print(f"    -> [Adapter] ✅ {file_name} 파일 저장 완료 (모든 서식 적용).")
            
            if not pivot_df_sorted.empty:
                 print(f"    -> [Adapter] 피벗 테이블 출력 샘플:\n{pivot_df_sorted.head()}")
            return True

        except Exception as e:
            print(f"    -> [Adapter] 🚨 엑셀 파일 쓰기 작업 중 예외 발생: {e}")
            return False # 파일 쓰기 실패