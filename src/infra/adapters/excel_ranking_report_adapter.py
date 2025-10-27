import pandas as pd
import datetime
from typing import Dict, Set, List, Optional, Any
import os

# (pip install openpyxl)
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
# [V5] 서식 관련 임포트
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import FILL_NONE

# 포트 임포트 (의존성)
from core.ports.excel_ranking_report_port import ExcelRankingReportPort

class ExcelRankingReportAdapter(ExcelRankingReportPort):
    """
    'ExcelRankingReportPort'의 구현체(Adapter).

    [V5 - 최종 완성본]
    'output' 폴더의 '2025일별수급순위정리표.xlsx' 파일을 열어,
    마지막 시트를 복사 -> 헤더 수정 -> 배경 초기화 -> 데이터 삽입 ->
    공통 항목 서식 적용 -> 자동 너비 맞춤 적용 후 저장합니다.
    (모든 단계를 작은 함수로 분리)

    # 엑셀 시트 레이아웃 가정 (필수):
    - A5: 날짜, B5: 요일
    - 데이터 영역: D5:L24
    - KOSPI 외국인: D5:D24 (종목명), E5:E24 (금액)
    - KOSPI 기관:   F5:F24 (종목명), G5:G24 (금액)
    - KOSDAQ 외국인: I5:I24 (종목명), J5:J24 (금액)
    - KOSDAQ 기관:   K5:K24 (종목명), L5:L24 (금액)
    """

    # --- 상수 정의 ---
    TOP_N_TO_PASTE = 20
    COMMON_STOCK_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid") # RGB(180, 198, 231)
    LAYOUT_MAP = {
        'KOSPI_foreigner': {'stock_col': 'D', 'value_col': 'E', 'start_row': 5, 'market': 'KOSPI'},
        'KOSPI_institutions': {'stock_col': 'F', 'value_col': 'G', 'start_row': 5, 'market': 'KOSPI'},
        'KOSDAQ_foreigner': {'stock_col': 'I', 'value_col': 'J', 'start_row': 5, 'market': 'KOSDAQ'},
        'KOSDAQ_institutions': {'stock_col': 'K', 'value_col': 'L', 'start_row': 5, 'market': 'KOSDAQ'},
    }
    DATA_RANGE_TO_CLEAR = "D5:L24" # 데이터 삽입 전 초기화할 영역
    COLUMNS_TO_AUTOFIT = ['D', 'F', 'I', 'K'] # 자동 너비 맞춤 적용할 종목명 열

    def __init__(self, base_path: str, file_name: str):
        self.output_path = base_path
        if not os.path.exists(self.output_path):
            os.makedirs(self.output_path)
        self.file_path = os.path.join(self.output_path, file_name)
        self.korean_weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        print(f"     -> [Adapter] ExcelRankingReportAdapter 초기화 (파일: {self.file_path})")

    # --- 워크북/시트 처리 함수 ---
    def _load_workbook(self) -> Optional[Workbook]:
        """엑셀 파일을 로드합니다."""
        try:
            book = openpyxl.load_workbook(self.file_path)
            print(f"     -> [Adapter] 파일 로드 성공: {self.file_path}")
            return book
        except FileNotFoundError:
            print(f"     -> [Adapter] 🚨 파일을 찾을 수 없습니다: {self.file_path}")
            return None
        except Exception as e:
            print(f"     -> [Adapter] 🚨 파일 로드 중 오류 발생: {e}")
            return None

    def _find_source_sheet(self, book: Workbook) -> Optional[Worksheet]:
        """워크북의 마지막 시트를 템플릿 원본으로 찾아 반환합니다."""
        if not book.sheetnames:
            print(f"     -> [Adapter] 🚨 파일에 시트가 하나도 없습니다.")
            return None
        source_sheet = book.worksheets[-1]
        print(f"     -> [Adapter] [Task 1] 원본 템플릿 시트 '{source_sheet.title}' 찾기 성공.")
        return source_sheet

    def _copy_and_prepare_sheet(
        self,
        book: Workbook,
        source_sheet: Worksheet,
        report_date: datetime.date
    ) -> Worksheet:
        """시트를 복사하고, 이름 설정, 중복 제거합니다."""
        new_sheet_name = report_date.strftime('%m%d')
        if new_sheet_name in book.sheetnames:
            print(f"     -> [Adapter] ⚠️ 기존 '{new_sheet_name}' 시트를 삭제합니다.")
            book.remove(book[new_sheet_name])
        new_sheet = book.copy_worksheet(source_sheet)
        new_sheet.title = new_sheet_name
        print(f"     -> [Adapter] [Task 2] 새 시트 '{new_sheet_name}' 생성 완료.")
        # 셀 크기 복사 로직 없음 (V3.4)
        return new_sheet

    def _update_sheet_headers(self, sheet: Worksheet, report_date: datetime.date):
        """새 시트의 A5(날짜)와 B5(요일) 셀을 업데이트합니다."""
        day_str = f"{report_date.day} 日"
        sheet['A5'] = day_str
        print(f"     -> [Adapter] [Task 3] A5 셀 수정 완료: {day_str}")
        weekday_str = self.korean_weekdays[report_date.weekday()]
        sheet['B5'] = weekday_str
        print(f"     -> [Adapter] [Task 4] B5 셀 수정 완료: {weekday_str}")

    # --- 데이터 처리 및 서식 함수 ---
    def _clear_data_area(self, ws: Worksheet):
        """지정된 데이터 영역의 값과 배경 서식을 초기화합니다."""
        print(f"     -> [Adapter] ... 데이터 영역 ({self.DATA_RANGE_TO_CLEAR}) 초기화 중...")
        for row in ws[self.DATA_RANGE_TO_CLEAR]:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=FILL_NONE)

    def _paste_single_list(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        layout: Dict[str, Any]
    ) -> int:
        """DataFrame의 상위 N개 데이터를 시트의 지정된 위치에 붙여넣습니다."""
        stock_col = layout['stock_col']
        value_col = layout['value_col']
        start_row = layout['start_row']

        # [V5] 이미 정렬된 DataFrame에서 상위 N개 가져오기 (head 사용)
        df_top_n = df.head(self.TOP_N_TO_PASTE)

        row_index = 0
        for _, row_data in df_top_n.iterrows():
            current_row = start_row + row_index
            ws[f"{stock_col}{current_row}"].value = row_data['종목명']
            ws[f"{value_col}{current_row}"].value = row_data['순매수_거래대금']
            row_index += 1
        return row_index # 실제로 붙여넣은 행의 수 반환

    def _apply_common_stock_format(
        self,
        ws: Worksheet,
        layout: Dict[str, Any],
        common_set: Set[str],
        pasted_rows_count: int
    ):
        """붙여넣은 데이터 중 공통 항목에 배경색 서식을 적용합니다."""
        stock_col = layout['stock_col']
        start_row = layout['start_row']

        for i in range(pasted_rows_count):
            current_row = start_row + i
            stock_cell = ws[f"{stock_col}{current_row}"]
            if stock_cell.value in common_set:
                stock_cell.fill = self.COMMON_STOCK_FILL

    def _clear_remaining_rows(
        self,
        ws: Worksheet,
        layout: Dict[str, Any],
        pasted_rows_count: int
    ):
        """데이터가 N개 미만일 경우, 템플릿의 남은 행을 지웁니다."""
        stock_col = layout['stock_col']
        value_col = layout['value_col']
        start_row = layout['start_row']

        for i in range(pasted_rows_count, self.TOP_N_TO_PASTE):
            current_row = start_row + i
            ws[f"{stock_col}{current_row}"].value = None
            ws[f"{value_col}{current_row}"].value = None
            ws[f"{stock_col}{current_row}"].fill = PatternFill(fill_type=FILL_NONE)

    def _apply_autofit(self, ws: Worksheet):
        """지정된 열에 자동 너비 맞춤(bestFit)을 적용합니다."""
        print(f"     -> [Adapter] ... 자동 너비 맞춤 적용 중 ({', '.join(self.COLUMNS_TO_AUTOFIT)} 열)...")
        for col_letter in self.COLUMNS_TO_AUTOFIT:
            ws.column_dimensions[col_letter].bestFit = True

    def _paste_and_format_data(
        self,
        ws: Worksheet,
        all_data: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ):
        """
        [V5] 데이터 삽입/서식 적용 워크플로우 오케스트레이션:
        배경 초기화 -> (각 리스트별) 데이터 삽입 -> 서식 적용 -> 남은 행 정리 -> 자동 너비 맞춤
        """
        self._clear_data_area(ws) # 1. 배경 초기화

        # 2. 4개 영역 순회하며 데이터 삽입 및 서식 적용
        for key, layout in self.LAYOUT_MAP.items():
            df = all_data.get(key)
            if df is None:
                print(f"     -> [Adapter] ⚠️ {key} 데이터가 없어 건너<0xEB><0x9B><0x81>니다.")
                continue

            market = layout['market']
            common_set = common_stocks.get(market, set())

            # 2.1 데이터 삽입
            pasted_count = self._paste_single_list(ws, df, layout)
            # 2.2 공통 항목 서식 적용
            self._apply_common_stock_format(ws, layout, common_set, pasted_count)
            # 2.3 남은 행 정리
            self._clear_remaining_rows(ws, layout, pasted_count)

            print(f"     -> [Adapter] ... {key} 영역 ({pasted_count}개) 처리 완료.")

        # 3. 자동 너비 맞춤 적용
        self._apply_autofit(ws)

    # --- 워크북 저장 함수 ---
    def _save_workbook(self, book: Workbook) -> bool:
        """워크북을 저장합니다."""
        try:
            book.save(self.file_path)
            print(f"     -> [Adapter] ✅ {self.file_path} 파일 저장 완료.")
            return True
        except Exception as e:
            print(f"     -> [Adapter] 🚨 파일 저장 중 오류 발생: {e}")
            return False

    # --- 메인 실행 함수 (Port 구현) ---
    def update_ranking_report(
        self,
        report_date: datetime.date,
        previous_date: datetime.date, # (사용 안 함)
        data_to_paste: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ) -> bool:
        """
        [V5] 전체 워크플로우 오케스트레이션:
        로드 -> 원본 찾기 -> 복사/준비 -> 헤더 업데이트 -> **데이터/서식/Autofit 적용** -> 저장
        """
        print(f"     -> [Adapter] 일별 수급 순위표 업데이트 시작 (파일: {self.file_path})")

        book = self._load_workbook()
        if book is None: return False

        source_sheet = self._find_source_sheet(book)
        if source_sheet is None: return False

        try:
            new_sheet = self._copy_and_prepare_sheet(book, source_sheet, report_date)
        except Exception as e:
             print(f"     -> [Adapter] 🚨 시트 복사/준비 중 오류 발생: {e}")
             return False

        try:
             self._update_sheet_headers(new_sheet, report_date)
        except Exception as e:
            print(f"     -> [Adapter] 🚨 헤더 업데이트 중 오류 발생: {e}")
            return False

        try:
            self._paste_and_format_data(new_sheet, data_to_paste, common_stocks)
        except Exception as e:
            print(f"     -> [Adapter] 🚨 데이터/서식 적용 중 오류 발생: {e}")
            return False

        return self._save_workbook(book)