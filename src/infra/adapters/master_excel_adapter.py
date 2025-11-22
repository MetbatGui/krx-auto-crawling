import pandas as pd
import datetime
from typing import Dict, List
import os

# (pip install openpyxl)
import openpyxl
# [V9/V13] dataframe_to_rows ì„í¬íŠ¸
from openpyxl.utils.dataframe import dataframe_to_rows
# [V16-V21] ì„œì‹ ì ìš©ì„ ìœ„í•œ Font, PatternFill ì„í¬íŠ¸
from openpyxl.styles import Font, PatternFill

from core.ports.master_report_port import MasterReportPort
from core.domain.models import KrxData

class MasterExcelAdapter(MasterReportPort):
    """
    MasterReportPortì˜ êµ¬í˜„ì²´(Adapter).
    ì›”ë³„ ëˆ„ì  ë° í”¼ë²— í…Œì´ë¸” ìƒì„±ì„ ë‹´ë‹¹í•©ë‹ˆë‹¤.
    """

    def __init__(self, base_path: str, file_name_prefix: str = "2025"):
        # ('output/ìˆœë§¤ìˆ˜ë„' í´ë” ìƒì„±)
        self.master_path = os.path.join(base_path, "ìˆœë§¤ìˆ˜ë„")
        if not os.path.exists(self.master_path):
            os.makedirs(self.master_path)
            
        # (íŒŒì¼ëª… í˜•ì‹: '...ìˆœë§¤ìˆ˜ë„(2025).xlsx')
        year_suffix = f"({file_name_prefix})"
        self.file_map: Dict[str, str] = {
            'KOSPI_foreigner': f'ì½”ìŠ¤í”¼ì™¸êµ­ì¸ìˆœë§¤ìˆ˜ë„{year_suffix}.xlsx',
            'KOSDAQ_foreigner': f'ì½”ìŠ¤ë‹¥ì™¸êµ­ì¸ìˆœë§¤ìˆ˜ë„{year_suffix}.xlsx',
            'KOSPI_institutions': f'ì½”ìŠ¤í”¼ê¸°ê´€ìˆœë§¤ìˆ˜ë„{year_suffix}.xlsx',
            'KOSDAQ_institutions': f'ì½”ìŠ¤ë‹¥ê¸°ê´€ìˆœë§¤ìˆ˜ë„{year_suffix}.xlsx',
        }

    def update_master_reports(self, data_list: List[KrxData]) -> None:
        """
        ìˆ˜ì§‘ëœ ë°ì´í„°ë¥¼ ë§ˆìŠ¤í„° íŒŒì¼ì— ëˆ„ì í•˜ê³  í”¼ë²— í…Œì´ë¸”ì„ ê°±ì‹ í•©ë‹ˆë‹¤.
        """
        for item in data_list:
            if item.data.empty:
                print(f"  [Adapter:MasterExcel] âš ï¸ {item.key} ë°ì´í„°ê°€ ë¹„ì–´ìˆì–´ ê±´ë„ˆëœë‹ˆë‹¤.")
                continue
            
            try:
                # ë‚ ì§œ ë¬¸ìì—´ì„ date ê°ì²´ë¡œ ë³€í™˜
                report_date = datetime.datetime.strptime(item.date_str, '%Y%m%d').date()
                self._update_single_report(item.key, item.data, report_date)
            except Exception as e:
                print(f"  [Adapter:MasterExcel] ğŸš¨ {item.key} ì—…ë°ì´íŠ¸ ì‹¤íŒ¨: {e}")

    def _update_single_report(
        self,
        report_key: str,
        daily_data: pd.DataFrame,
        report_date: datetime.date
    ) -> bool:
        
        file_name = self.file_map.get(report_key)
        if not file_name:
            print(f"    -> [Adapter:MasterExcel] ğŸš¨ '{report_key}'ì— í•´ë‹¹í•˜ëŠ” íŒŒì¼ëª…ì„ ëª¨ë¦…ë‹ˆë‹¤.")
            return False

        file_path = os.path.join(self.master_path, file_name)
        
        # [V13] ì‹œíŠ¸ ì´ë¦„ ì •ì˜
        sheet_name = report_date.strftime('%b').upper()
        pivot_sheet_name = report_date.strftime('%m%d') 
        
        date_str = report_date.strftime('%Y%m%d')
        date_int = int(date_str) 

        print(f"    -> [Adapter:MasterExcel] {file_name} íŒŒì¼ ì—…ë°ì´íŠ¸ ì‹œì‘...")

        # --- [V22] ë¹ ë¥¸ ê±´ë„ˆë›°ê¸° ë¡œì§ ---
        # 1. íŒŒì¼ì„ ì—´ê¸° ì „ì—, íŒŒì¼ì´ ì¡´ì¬í•˜ëŠ”ì§€ í™•ì¸
        if os.path.exists(file_path):
            try:
                # 2. ì¡´ì¬í•œë‹¤ë©´, ì‹œíŠ¸ ì´ë¦„ë§Œ ë¹ ë¥´ê²Œ ì½ì–´ì˜´ (read_only=True)
                #    (íŒŒì¼ì´ ê¹¨ì¡Œì„ ê²½ìš°ë¥¼ ëŒ€ë¹„í•´ try/except)
                book = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = book.sheetnames
                book.close()
                
                # 3. ì˜¤ëŠ˜ ë‚ ì§œì˜ í”¼ë²— ì‹œíŠ¸ê°€ ì´ë¯¸ ìˆëŠ”ì§€ í™•ì¸
                if pivot_sheet_name in sheet_names:
                    print(f"    -> [Adapter:MasterExcel] âš ï¸ '{pivot_sheet_name}' í”¼ë²— ì‹œíŠ¸ê°€ ì´ë¯¸ ì¡´ì¬í•˜ì—¬ [ë¹ ë¥¸ ê±´ë„ˆë›°ê¸°]ë¥¼ ì‹¤í–‰í•©ë‹ˆë‹¤.")
                    return True # Trueë¥¼ ë°˜í™˜í•˜ì—¬ íŒŒì´í”„ë¼ì¸ ê³„ì† ì§„í–‰
            except Exception as e:
                # (ì˜ˆ: íŒŒì¼ì´ ê¹¨ì¡Œê±°ë‚˜, zip íŒŒì¼ì´ ì•„ë‹Œ ê²½ìš°)
                print(f"    -> [Adapter:MasterExcel] âš ï¸ ë¹ ë¥¸ ê±´ë„ˆë›°ê¸° ê²€ì‚¬ ì¤‘ íŒŒì¼({file_name})ì„ ì½ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {e}")
                print(f"    -> [Adapter:MasterExcel] âš ï¸ (íŒŒì¼ì„ ë®ì–´ì“°ê¸° ìœ„í•´, ì „ì²´ ë¡œì§ì„ ê³„ì† ì§„í–‰í•©ë‹ˆë‹¤.)")
                pass # ì—ëŸ¬ê°€ ë‚¬ìœ¼ë¯€ë¡œ, ì •ìƒ ë¡œì§ì„ íƒœì›Œì„œ ë®ì–´ì“°ë„ë¡ ìœ ë„
        # --- [V22] ë¹ ë¥¸ ê±´ë„ˆë›°ê¸° ë ---


        # --- (V22) (ì—¬ê¸°ê¹Œì§€ ì™”ë‹¤ë©´, í”¼ë²— ì‹œíŠ¸ê°€ ì—†ê±°ë‚˜ íŒŒì¼ì´ ì—†ìœ¼ë¯€ë¡œ, V21ì˜ ì „ì²´ ë¡œì§ì„ ì‹¤í–‰) ---


        # --- 1. [V11] ìƒˆ ë°ì´í„°ë¥¼ ì—‘ì…€ ìŠ¤í‚¤ë§ˆë¡œ ë²ˆì—­ ---
        try:
            data_dict = {
                'ì¼ì': date_int, 
                'ì¢…ëª©': daily_data['ì¢…ëª©ëª…'],
                'ê¸ˆì•¡': pd.to_numeric(daily_data['ìˆœë§¤ìˆ˜_ê±°ë˜ëŒ€ê¸ˆ'])
            }
            new_data_formatted = pd.DataFrame(data_dict)
            new_data_formatted = new_data_formatted[['ì¼ì', 'ì¢…ëª©', 'ê¸ˆì•¡']]

        except KeyError as e:
            print(f"    -> [Adapter:MasterExcel] ğŸš¨ 'daily_data'ì— í•„ìš”í•œ ì»¬ëŸ¼ì´ ì—†ìŠµë‹ˆë‹¤: {e}")
            return False

        # --- 2. [V15] ê¸°ì¡´ ë°ì´í„° ì½ê¸° (Pandas, ì¤‘ë³µ ê²€ì‚¬ìš© - 1íšŒ ì½ê¸°) ---
        excel_columns = ['ì¼ì', 'ì¢…ëª©', 'ê¸ˆì•¡']
        existing_df = pd.DataFrame(columns=excel_columns)
        sheet_exists = False 
        try:
            read_df = pd.read_excel(
                file_path, 
                sheet_name=sheet_name, 
                engine='openpyxl', 
                skiprows=1,
                dtype={'ì¼ì': int}
            )
            sheet_exists = True 
            if not read_df.empty:
                if all(col in read_df.columns for col in read_df.columns):
                        existing_df = read_df[excel_columns].copy()
                else:
                    print(f"    -> [Adapter:MasterExcel] âš ï¸ {sheet_name} ì‹œíŠ¸ í—¤ë”ê°€ ê¹¨ì ¸ ì½ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                    existing_df = pd.DataFrame(columns=excel_columns)
            print(f"    -> [Adapter:MasterExcel] ê¸°ì¡´ '{sheet_name}' ì‹œíŠ¸ ë°ì´í„° ({len(existing_df)}ì¤„) ë¡œë“œ ì™„ë£Œ.")
        except FileNotFoundError:
            print(f"    -> [Adapter:MasterExcel] âš ï¸ ìƒˆ íŒŒì¼ '{file_name}'ì´ ìƒì„±ë©ë‹ˆë‹¤.")
        except (ValueError, KeyError) as e:
            print(f"    -> [Adapter:MasterExcel] âš ï¸ íŒŒì¼ì€ ìˆìœ¼ë‚˜ '{sheet_name}' ì‹œíŠ¸ê°€ ì—†ì–´ ìƒˆë¡œ ìƒì„±í•©ë‹ˆë‹¤.")
        except Exception as e:
            print(f"    -> [Adapter:MasterExcel] ğŸš¨ íŒŒì¼ ë¡œë“œ ì¤‘ ì˜ˆìƒì¹˜ ëª»í•œ ì˜¤ë¥˜: {e}")
            return False

        # --- 3. [V13 ìˆ˜ì •] ì¤‘ë³µ ë‚ ì§œ ê²€ì‚¬ ---
        if date_int in existing_df['ì¼ì'].values: 
            print(f"    -> [Adapter:MasterExcel] âš ï¸ {date_int} ë°ì´í„°ê°€ '{sheet_name}'ì— ì´ë¯¸ ì¡´ì¬í•˜ì—¬ ë¬´ì‹œí•©ë‹ˆë‹¤.")
            new_data_formatted = pd.DataFrame()
            print("         (ë°ì´í„° ì¶”ê°€ëŠ” ê±´ë„ˆë›°ê³ , í”¼ë²— í…Œì´ë¸” ìƒì„±(2ë‹¨ê³„)ì€ ì§„í–‰í•©ë‹ˆë‹¤.)")
        
        if not new_data_formatted.empty:
            print(f"    -> [Adapter:MasterExcel] ìƒˆ ë°ì´í„° ({len(new_data_formatted)}ì¤„) ì¶”ê°€ ì¤€ë¹„...")

        # --- 4. [V15] í”¼ë²— ìƒì„±ì„ ìœ„í•´ ë©”ëª¨ë¦¬ì—ì„œ ì „ì²´ ë°ì´í„° ì¤€ë¹„ ---
        print(f"    -> [Adapter:MasterExcel] ë©”ëª¨ë¦¬ì—ì„œ í”¼ë²—ìš© ì „ì²´ ë°ì´í„° ì¤€ë¹„...")
        if not new_data_formatted.empty:
            full_data_df = pd.concat([existing_df, new_data_formatted], ignore_index=True)
        else:
            full_data_df = existing_df.copy()

        # --- 5. [V15] í”¼ë²— í…Œì´ë¸” ê³„ì‚° (íŒŒì¼ ì“°ê¸° ì „) ---
        print(f"    -> [Adapter:MasterExcel] '{pivot_sheet_name}' í”¼ë²— í…Œì´ë¸” ê³„ì‚° ì‹œì‘...")
        pivot_df_sorted = pd.DataFrame()
        # í”¼ë²—ì„ ë§Œë“¤ê¸° ì „, 'ê¸ˆì•¡' ì»¬ëŸ¼ì„ ìˆ«ìë¡œ ê°•ì œ ë³€í™˜í•©ë‹ˆë‹¤.
        # (ê¸°ì¡´ ë°ì´í„°ê°€ "1,234,000" ì²˜ëŸ¼ ë¬¸ìì—´ë¡œ ë¡œë“œë˜ì—ˆì„ ê²½ìš° ëŒ€ë¹„)
        if not full_data_df.empty:
            try:
                # 1. ì‰¼í‘œ ë“± ë¶ˆí•„ìš”í•œ ë¬¸ì ì œê±° (ìˆ«ì, ì†Œìˆ˜ì , ë§ˆì´ë„ˆìŠ¤ ë¶€í˜¸ ì™¸)
                full_data_df['ê¸ˆì•¡'] = full_data_df['ê¸ˆì•¡'].astype(str).str.replace(r'[^0-9.-]', '', regex=True)
                # 2. ë¹ˆ ë¬¸ìì—´ì€ 0ìœ¼ë¡œ
                full_data_df['ê¸ˆì•¡'] = full_data_df['ê¸ˆì•¡'].replace('', 0)
                # 3. ìˆ«ìë¡œ ë³€í™˜ (ì˜¤ë¥˜ ì‹œ NaN)
                full_data_df['ê¸ˆì•¡'] = pd.to_numeric(full_data_df['ê¸ˆì•¡'], errors='coerce')
                # 4. NaNì„ 0ìœ¼ë¡œ (ê²°ì¸¡ì¹˜ ë°©ì§€)
                full_data_df['ê¸ˆì•¡'] = full_data_df['ê¸ˆì•¡'].fillna(0)
            except Exception as clean_e:
                print(f" Â  Â -> [Adapter:MasterExcel] ğŸš¨ 'ê¸ˆì•¡' ì»¬ëŸ¼ ìˆ«ì ë³€í™˜ ì¤‘ ì˜¤ë¥˜: {clean_e}")
                # (ì˜¤ë¥˜ê°€ ë‚˜ë„ ì¼ë‹¨ ì§„í–‰ ì‹œë„)
        # --- [ìˆ˜ì • ì½”ë“œ ë] ---
        if full_data_df.empty:
             print(f"    -> [Adapter:MasterExcel] âš ï¸ '{sheet_name}' ì›ë³¸ ë°ì´í„°ê°€ ë¹„ì–´ìˆì–´ í”¼ë²—ì„ ìƒì„±í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        else:
            try:
                # [V21] 'ì´ê³„' ì¶”ê°€ ì „ì˜ ì›ë³¸ í”¼ë²— (ì˜¤ëŠ˜ ë‚ ì§œ ì—´ ì°¾ê¸°ìš©)
                pivot_df = pd.pivot_table(
                    full_data_df,
                    values='ê¸ˆì•¡',
                    index='ì¢…ëª©',
                    columns='ì¼ì',
                    aggfunc='sum'
                )
                pivot_df['ì´ê³„'] = pivot_df.sum(axis=1)
                pivot_df_sorted = pivot_df.sort_values(by='ì´ê³„', ascending=False)
                print(f"    -> [Adapter:MasterExcel] í”¼ë²— í…Œì´ë¸” ê³„ì‚° ì™„ë£Œ.")
            except Exception as e:
                print(f"    -> [Adapter:MasterExcel] ğŸš¨ í”¼ë²— í…Œì´ë¸” ê³„ì‚° ì¤‘ ì˜ˆì™¸ ë°œìƒ: {e}")
                return False 

        # --- 6. [V22] ì—‘ì…€ íŒŒì¼ í•œ ë²ˆì— ì“°ê¸° (ëª¨ë“  ì„œì‹/ìˆœì„œ ì ìš©) ---
        print(f"    -> [Adapter:MasterExcel] ì—‘ì…€ íŒŒì¼ ì“°ê¸° ì‘ì—… ì‹œì‘ ({file_name})...")
        try:
            try:
                book = openpyxl.load_workbook(file_path)
            except FileNotFoundError:
                book = openpyxl.Workbook()
                if 'Sheet' in book.sheetnames:
                    book.remove(book['Sheet'])
            
            # (V22) "ë¹ ë¥¸ ê±´ë„ˆë›°ê¸°"ë¥¼ í†µê³¼í–ˆìœ¼ë¯€ë¡œ, ì—¬ê¸°ì„œëŠ” V21-Skipì˜ ì¤‘ë³µ ê²€ì‚¬ ë¡œì§ì´ í•„ìš” ì—†ìŒ

            # --- [ì“°ê¸° 1ë‹¨ê³„: 'OCT' ì‹œíŠ¸ ëˆ„ì  (V19 ìˆœì„œ ë³´ì¥)] ---
            if not new_data_formatted.empty:
                if sheet_exists: 
                    ws = book[sheet_name]
                    print(f"        -> [1ë‹¨ê³„] '{sheet_name}' ì‹œíŠ¸ ë§ˆì§€ë§‰ í–‰({ws.max_row})ì— ì¶”ê°€í•©ë‹ˆë‹¤.")
                    for row in dataframe_to_rows(new_data_formatted, index=False, header=False):
                        ws.append(row)
                else:
                    # [V19] 'OCT' ì‹œíŠ¸ê°€ ì—†ìœ¼ë©´, ë§ˆì§€ë§‰ ì‹œíŠ¸(ì´ê²°ì‚°) "ì•ì—" ìƒì„±
                    data_sheet_index = 0
                    if len(book.sheetnames) > 0:
                        data_sheet_index = len(book.sheetnames) - 1 
                    
                    ws = book.create_sheet(title=sheet_name, index=data_sheet_index)
                    print(f"        -> [1ë‹¨ê³„] ìƒˆ '{sheet_name}' ì‹œíŠ¸ë¥¼ {data_sheet_index}ë²ˆì§¸ (ì´ê²°ì‚° ì•)ì— ìƒì„±.")
                    ws.append([]) # A1
                    ws.append(list(new_data_formatted.columns)) # A2
                    for row in dataframe_to_rows(new_data_formatted, index=False, header=False):
                        ws.append(row)
                print(f"        -> [1ë‹¨ê³„] âœ… '{sheet_name}' ì‹œíŠ¸ ëˆ„ì  ì™„ë£Œ (ì•„ì§ ì €ì¥ ì „).")
            else:
                print(f"        -> [1ë‹¨ê³„] â­ï¸ ë°ì´í„°ê°€ (ì¤‘ë³µ ë“±ìœ¼ë¡œ) ë¹„ì–´ìˆì–´ ëˆ„ì ì„ ê±´ë„ˆëœë‹ˆë‹¤.")

            # --- [ì“°ê¸° 2ë‹¨ê³„: '1023' í”¼ë²— ë®ì–´ì“°ê¸° ë° ì„œì‹ ì ìš© (V22)] ---
            
            # (V22) V21-Skipê³¼ ë‹¬ë¦¬, í”¼ë²— ì‹œíŠ¸ê°€ *í˜¹ì‹œë¼ë„* ì¡´ì¬í•˜ë©´ (ì˜ˆ: ê¹¨ì§„ íŒŒì¼ ë³µêµ¬ ì‹œ)
            # ë®ì–´ì“°ê¸° ìœ„í•´ "ì‚­ì œ" ë¡œì§ì„ ë‹¤ì‹œ ë³µì›
            if pivot_sheet_name in book.sheetnames:
                print(f"        -> [2ë‹¨ê³„] âš ï¸ (ê²½ê³ ) '{pivot_sheet_name}' ì‹œíŠ¸ê°€ ì¡´ì¬í•˜ì—¬ ë®ì–´ì”ë‹ˆë‹¤.")
                book.remove(book[pivot_sheet_name])
            
            # [V19] 'OCT' ì‹œíŠ¸ "ë°”ë¡œ ì•"ì— í”¼ë²— ì‹œíŠ¸ ìƒì„±
            try:
                data_sheet_index = book.sheetnames.index(sheet_name)
            except ValueError:
                if not new_data_formatted.empty: 
                    data_sheet_index = book.sheetnames.index(sheet_name)
                else: 
                    print(f"        -> [2ë‹¨ê³„] ğŸš¨ ë²„ê·¸: '{sheet_name}' ì‹œíŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. í”¼ë²—ì„ ë§¨ ë’¤ì— ìƒì„±í•©ë‹ˆë‹¤.")
                    data_sheet_index = -1 
                
            pivot_ws = book.create_sheet(title=pivot_sheet_name, index=data_sheet_index)
            print(f"        -> [2ë‹¨ê³„] '{pivot_sheet_name}' ì‹œíŠ¸ë¥¼ {data_sheet_index}ë²ˆì§¸ ('{sheet_name}' ì•)ì— ìƒì„±.")

            
            # [V21] 1. ì„œì‹ ìŠ¤íƒ€ì¼ ì •ì˜
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # í•˜ëŠ˜ìƒ‰
            red_font = Font(color="FF0000") # ë¹¨ê°„ìƒ‰
            top_5_colors = ["FF0000", "FFC000", "FFFF00", "92D050", "00B0F0"]
            top_5_fills = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in top_5_colors]
            
            # [V20] 2. 2í–‰ ì•„ë˜ë¡œ ë‚´ë¦¬ê¸° (A1, A2ëŠ” ë¹„ì›€)
            pivot_ws.append([]) # ë¹ˆ í–‰ (1í–‰)
            pivot_ws.append([]) # ë¹ˆ í–‰ (2í–‰)
            
            # [V20] 3. Aì—´ ë„ˆë¹„ ì¡°ì • (165px -> 22.86)
            pivot_ws.column_dimensions['A'].width = 22.86
            
            # [V20] 4. í”¼ë²— ë°ì´í„° ì“°ê¸° (A3ë¶€í„°) ë° ìµœëŒ€ ì—´ ê³„ì‚°
            max_col_to_style = 1 
            data_start_row = 5 
            
            if not pivot_df_sorted.empty:
                print(f"        -> [2ë‹¨ê³„] '{pivot_sheet_name}' ì‹œíŠ¸ A3ì…€ë¶€í„° í”¼ë²— ë°ì´í„° ì“°ê¸°...")
                max_col_to_style = 1 + len(pivot_df_sorted.columns) 
                
                for r in dataframe_to_rows(pivot_df_sorted, index=True, header=True):
                    pivot_ws.append(r) # A3ë¶€í„° ì¨ì§
                print(f"        -> [2ë‹¨ê³„] âœ… '{pivot_sheet_name}' ì‹œíŠ¸ ë®ì–´ì“°ê¸° ì™„ë£Œ (ì•„ì§ ì €ì¥ ì „).")
            else:
                 print(f"        -> [2ë‹¨ê³„] âš ï¸ í”¼ë²— ë°ì´í„°ê°€ ë¹„ì–´ìˆì–´ '{pivot_sheet_name}'ëŠ” ë¹ˆ ì‹œíŠ¸ë¡œ ìƒì„±ë©ë‹ˆë‹¤.")

            # [V20] 5. A3:?4 (ë™ì ) ë²”ìœ„ì— ë°°ê²½ìƒ‰ ì ìš© (3í–‰, 4í–‰)
            print(f"        -> [2ë‹¨ê³„] í—¤ë” ì˜ì—­ (3-4í–‰, {max_col_to_style}ê°œ ì—´) ë°°ê²½ìƒ‰ ì ìš©...")
            for row in pivot_ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=max_col_to_style):
                for cell in row:
                    cell.fill = header_fill

            # [V21] 6. Top 20 'ì´ê³„' ì¢…ëª© (Aì—´) í°íŠ¸ ë¹¨ê°„ìƒ‰ ì ìš©
            if not pivot_df_sorted.empty:
                top_20_end_row = (data_start_row - 1) + 20
                print(f"        -> [2ë‹¨ê³„] Top 20 (A{data_start_row}:A{top_20_end_row}) ì¢…ëª© í°íŠ¸ ì ìš©...")
                safe_end_row = min(top_20_end_row, pivot_ws.max_row) 
                for row in pivot_ws.iter_rows(min_row=data_start_row, max_row=safe_end_row, min_col=1, max_col=1):
                    row[0].font = red_font
            
            # [V21] 7. ë‹¹ì¼ Top 5 ë°°ê²½ìƒ‰ ì ìš©
            if not pivot_df_sorted.empty and (date_int in pivot_df.columns):
                try:
                    date_col_idx_in_df = list(pivot_df.columns).index(date_int)
                    target_col_in_ws = date_col_idx_in_df + 2 
                    top_5_series = pivot_df_sorted[date_int].nlargest(5)
                    top_5_series = top_5_series[top_5_series > 0] 
                    top_5_stock_map = {stock_name: fill for stock_name, fill in zip(top_5_series.index, top_5_fills)}

                    if top_5_stock_map:
                        print(f"        -> [2ë‹¨ê³„] ë‹¹ì¼ Top {len(top_5_stock_map)} ({date_int}, {target_col_in_ws}ì—´) ë°°ê²½ìƒ‰ ì ìš©...")
                        for row in pivot_ws.iter_rows(min_row=data_start_row, max_row=pivot_ws.max_row, min_col=1, max_col=target_col_in_ws):
                            stock_name_cell = row[0] 
                            if stock_name_cell.value in top_5_stock_map:
                                target_cell = row[target_col_in_ws - 1] 
                                target_cell.fill = top_5_stock_map[stock_name_cell.value]
                    else:
                        print(f"        -> [2ë‹¨ê³„] âš ï¸ ë‹¹ì¼ Top 5 (ì–‘ìˆ˜) ë°ì´í„°ê°€ ì—†ì–´ ë°°ê²½ìƒ‰ì„ ê±´ë„ˆëœë‹ˆë‹¤.")
                except ValueError:
                     print(f"        -> [2ë‹¨ê³„] âš ï¸ ì˜¤ëŠ˜ ë‚ ì§œ({date_int})ë¥¼ í”¼ë²— ì»¬ëŸ¼ì—ì„œ ì°¾ì„ ìˆ˜ ì—†ì–´ ë°°ê²½ìƒ‰ì„ ê±´ë„ˆëœë‹ˆë‹¤.")
            else:
                 print(f"        -> [2ë‹¨ê³„] âš ï¸ í”¼ë²—ì´ ë¹„ì–´ìˆê±°ë‚˜ ì˜¤ëŠ˜ ë‚ ì§œ({date_int})ê°€ ì—†ì–´ ë°°ê²½ìƒ‰ì„ ê±´ë„ˆëœë‹ˆë‹¤.")
                 
            # [ìµœì¢… ì €ì¥] - ëª¨ë“  ë³€ê²½ì‚¬í•­ì„ í•œ ë²ˆì— ì €ì¥
            book.save(file_path)
            print(f"    -> [Adapter:MasterExcel] âœ… {file_name} íŒŒì¼ ì €ì¥ ì™„ë£Œ (ëª¨ë“  ì„œì‹ ì ìš©).")
            
            if not pivot_df_sorted.empty:
                 print(f"    -> [Adapter:MasterExcel] í”¼ë²— í…Œì´ë¸” ì¶œë ¥ ìƒ˜í”Œ:\n{pivot_df_sorted.head()}")
            return True

        except Exception as e:
            print(f"    -> [Adapter:MasterExcel] ğŸš¨ ì—‘ì…€ íŒŒì¼ ì“°ê¸° ì‘ì—… ì¤‘ ì˜ˆì™¸ ë°œìƒ: {e}")
            return False # íŒŒì¼ ì“°ê¸° ì‹¤íŒ¨