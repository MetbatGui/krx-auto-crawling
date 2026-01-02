"""순위표 Excel 리포트 어댑터"""

import datetime
import pandas as pd
from typing import Dict, Set
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill

from core.ports.ranking_report_port import RankingReportPort
from core.ports.storage_port import StoragePort
from infra.adapters.excel.excel_formatter import ExcelFormatter
from infra.adapters.excel.excel_sheet_builder import ExcelSheetBuilder


class RankingExcelAdapter(RankingReportPort):
    """순위표를 Excel 형식으로 생성하는 어댑터.

    RankingReportPort 인터페이스의 Excel 구현체입니다.
    ExcelFormatter와 ExcelSheetBuilder 유틸리티를 조합하여 사용합니다.

    Attributes:
        storage (StoragePort): 파일 저장/로드 포트.
        file_path (str): Excel 파일 경로.
    """
    
    TOP_N = 30
    LAYOUT_MAP = {
        'KOSPI_foreigner': {'stock_col': 'E', 'value_col': 'F', 'start_row': 5, 'market': 'KOSPI'},
        'KOSPI_institutions': {'stock_col': 'H', 'value_col': 'I', 'start_row': 5, 'market': 'KOSPI'},
        'KOSDAQ_foreigner': {'stock_col': 'L', 'value_col': 'M', 'start_row': 5, 'market': 'KOSDAQ'},
        'KOSDAQ_institutions': {'stock_col': 'O', 'value_col': 'P', 'start_row': 5, 'market': 'KOSDAQ'},
    }
    # Top 30 기준 Clear Range: 5행부터 34행까지 (30개)
    DATA_RANGE_TO_CLEAR = "E5:P34"
    COLUMNS_TO_AUTOFIT = ['E', 'H', 'L', 'O']
    KOREAN_WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]
    
    # 기본 템플릿 경로 상수 (StorageRoot 기준)
    DEFAULT_TEMPLATE_PATH = "template/template_일별수급순위정리표.xlsx"
    
    def __init__(
        self, 
        source_storage: StoragePort, 
        target_storages: List[StoragePort], 
        file_name: str = "2025일별수급순위정리표.xlsx",
        template_file_path: str = None
    ):
        """RankingExcelAdapter 초기화.

        Args:
            source_storage (StoragePort): 파일을 로드할 저장소 (예: GoogleDriveAdapter).
            target_storages (List[StoragePort]): 파일을 저장할 저장소 리스트 (예: [LocalStorageAdapter, GoogleDriveAdapter]).
            file_name (str): Excel 파일명.
            template_file_path (str): 템플릿 파일 경로 (Optional).
        """
        self.source_storage = source_storage
        self.target_storages = target_storages
        self.file_path = file_name
        self.template_file_path = template_file_path or self.DEFAULT_TEMPLATE_PATH
        
        print(f"[Adapter:RankingExcel] 초기화 완료 (파일: {self.file_path}, 템플릿: {self.template_file_path})")
    
    def update_report(
        self,
        report_date: datetime.date,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ) -> bool:
        """순위표 리포트를 업데이트합니다.

        Args:
            report_date (datetime.date): 리포트 날짜.
            data_map (Dict[str, pd.DataFrame]): 데이터 딕셔너리.
            common_stocks (Dict[str, Set[str]]): 공통 종목 딕셔너리.

        Returns:
            bool: 성공 여부.
        """
        book = self._load_workbook()
        if not book:
            return False
        
        new_sheet = self._create_new_sheet(book, report_date)
        if not new_sheet:
            return False
        
        self._update_sheet_content(new_sheet, report_date, data_map, common_stocks)
        
        return self._save_workbook(book)
    
    def _load_workbook(self) -> Workbook | None:
        """워크북을 로드합니다. 파일이 없으면 템플릿을 복사하여 시작합니다."""
        print(f"    -> [Adapter:RankingExcel] 로드 시도 ({self.source_storage.__class__.__name__})...")
        
        # 파일이 존재하는지 확인
        if not self.source_storage.path_exists(self.file_path):
            print(f"    -> [Adapter:RankingExcel] 파일이 없어 템플릿 복사를 시도합니다: {self.template_file_path}")
            
            # 템플릿 파일 로드 (바이트)
            template_data = self.source_storage.get_file(self.template_file_path)
            if template_data:
                # 타겟 경로에 템플릿 저장 (Source Storage에 우선 저장)
                # 주의: 로드는 source_storage에서 하므로, source_storage에 파일이 있어야 함.
                # 보통 source_storage는 로컬이거나 공유 드라이브일 것임.
                if self.source_storage.put_file(self.file_path, template_data):
                    print(f"    -> [Adapter:RankingExcel] 템플릿 복사 성공")
                else:
                    print(f"    -> [Adapter:RankingExcel] 🚨 템플릿 저장 실패")
                    return None
            else:
                print(f"    -> [Adapter:RankingExcel] 🚨 템플릿 파일을 찾을 수 없습니다: {self.template_file_path}")
                # 템플릿이 없으면 새 파일 생성 로직으로 갈 수도 있지만, 여기서는 실패 처리
                return None

        # 파일 로드
        book = self.source_storage.load_workbook(self.file_path)
        if not book:
            print(f"    -> [Adapter:RankingExcel] 🚨 워크북 로드 실패: {self.file_path}")
            return None
            
        return book
    
    def _create_new_sheet(self, book: Workbook, report_date: datetime.date) -> Worksheet | None:
        """새로운 시트를 생성합니다 (템플릿 시트 복제)."""
        try:
            sheet_name = report_date.strftime('%m%d')
            
            # 이미 시트가 있으면 삭제
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            
            # 복제 소스 시트 결정 ('template' 시트 우선)
            if 'template' in book.sheetnames:
                source_sheet = book['template']
                print(f"    -> [Adapter:RankingExcel] 'template' 시트 복제 사용")
            else:
                source_sheet = book.worksheets[-1]
                print(f"    -> [Adapter:RankingExcel] 'template' 시트가 없어 마지막 시트 복제 사용")
            
            new_sheet = book.copy_worksheet(source_sheet)
            new_sheet.title = sheet_name
            
            
            
            print(f"    -> [Adapter:RankingExcel] '{sheet_name}' 시트 생성 완료")
            return new_sheet
        except Exception as e:
            print(f"    -> [Adapter:RankingExcel] 🚨 시트 생성 실패: {e}")
            return None
    
    def _update_sheet_content(
        self,
        sheet: Worksheet,
        report_date: datetime.date,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ):
        """시트 내용을 업데이트합니다."""
        self._update_headers(sheet, report_date)
        self._clear_data_area(sheet)
        self._paste_data_and_apply_format(sheet, data_map, common_stocks)
        self._apply_autofit(sheet)
    
    def _update_headers(self, sheet: Worksheet, report_date: datetime.date):
        """헤더를 업데이트합니다.
        
        A3: 월 (예: "11 月")
        A5: 일 (예: "21 日")
        B5: 요일 (예: "금")
        """
        sheet['A3'] = f"{report_date.month} 月"
        sheet['A5'] = f"{report_date.day} 日"
        sheet['B5'] = self.KOREAN_WEEKDAYS[report_date.weekday()]
    
    def _clear_data_area(self, sheet: Worksheet):
        """데이터 영역을 초기화합니다."""
        for row in sheet[self.DATA_RANGE_TO_CLEAR]:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill()
    
    def _paste_data_and_apply_format(
        self,
        sheet: Worksheet,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ):
        """데이터를 붙여넣고 서식을 적용합니다."""
        for key, layout in self.LAYOUT_MAP.items():
            df = data_map.get(key)
            if df is None or df.empty:
                continue
            
            pasted_count = ExcelSheetBuilder.paste_ranking_data(sheet, df, layout, self.TOP_N)
            
            market = layout['market']
            if market in common_stocks:
                ExcelFormatter.apply_common_stock_fill(
                    sheet,
                    layout['stock_col'],
                    layout['start_row'],
                    pasted_count,
                    common_stocks[market]
                )
            
            ExcelSheetBuilder.clear_ranking_remaining_rows(sheet, layout, pasted_count, self.TOP_N)
    
    def _apply_autofit(self, sheet: Worksheet):
        """열 너비를 자동 조정합니다."""
        for col in self.COLUMNS_TO_AUTOFIT:
            sheet.column_dimensions[col].bestFit = True
            sheet.column_dimensions[col].auto_size = True
    
    def _save_workbook(self, book: Workbook) -> bool:
        """워크북을 저장합니다 (Target Storages 사용)."""
        all_success = True
        for storage in self.target_storages:
            success = storage.save_workbook(book, self.file_path)
            if success:
                print(f"    -> [Adapter:RankingExcel] ✅ {storage.__class__.__name__} 순위표 저장 완료")
            else:
                all_success = False
        return all_success
