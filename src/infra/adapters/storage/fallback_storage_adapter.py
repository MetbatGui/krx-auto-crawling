from typing import Optional, Any
import pandas as pd
import openpyxl
from core.ports.storage_port import StoragePort

class FallbackStorageAdapter(StoragePort):
    """두 개의 저장소를 순차적으로 조회하는 어댑터 (읽기 전용 Fallback).
    
    Primary 저장소에서 파일을 찾지 못하면 Secondary 저장소를 조회합니다.
    쓰기 작업(Save)은 Primary에만 수행하거나, 안전을 위해 에러를 발생시킬 수 있습니다.
    (현재 구조에서는 Source 용도로만 사용되므로 읽기 위주입니다.)
    
    Attributes:
        primary (StoragePort): 주 저장소.
        secondary (StoragePort): 보조 저장소 (Fallback).
    """

    def __init__(self, primary: StoragePort, secondary: StoragePort):
        self.primary = primary
        self.secondary = secondary

    def path_exists(self, path: str) -> bool:
        """Primary 확인 후 없으면 Secondary 확인합니다.
        
        Args:
            path (str): 확인할 경로.
            
        Returns:
            bool: 존재 여부.
        """
        if self.primary.path_exists(path):
            return True
        return self.secondary.path_exists(path)

    def load_workbook(self, path: str) -> Optional[openpyxl.Workbook]:
        """Primary에서 로드 시도, 실패 시 Secondary에서 로드합니다.
        
        Args:
            path (str): 로드할 파일 경로.
            
        Returns:
            Optional[openpyxl.Workbook]: 로드된 Workbook 객체, 실패 시 None.
        """
        # Primary 시도
        if self.primary.path_exists(path):
            wb = self.primary.load_workbook(path)
            if wb:
                print(f"[FallbackStorage] Primary({self.primary.__class__.__name__})에서 로드 성공: {path}")
                return wb
        
        # Secondary 시도
        if self.secondary.path_exists(path):
            print(f"[FallbackStorage] Primary에 없음. Secondary({self.secondary.__class__.__name__})에서 로드 시도: {path}")
            wb = self.secondary.load_workbook(path)
            if wb:
                return wb
                
        return None

    def load_dataframe(self, path: str, sheet_name: str = None, **kwargs) -> pd.DataFrame:
        """Primary에서 로드 시도, 실패 시 Secondary에서 로드합니다.
        
        Args:
            path (str): 로드할 파일 경로.
            sheet_name (str, optional): 시트 이름.
            **kwargs: 추가 옵션.
            
        Returns:
            pd.DataFrame: 로드된 DataFrame.
        """
        # Primary 시도
        if self.primary.path_exists(path):
            df = self.primary.load_dataframe(path, sheet_name, **kwargs)
            if not df.empty:
                print(f"[FallbackStorage] Primary({self.primary.__class__.__name__})에서 DataFrame 로드 성공: {path}")
                return df
        
        # Secondary 시도
        print(f"[FallbackStorage] Primary에 없음. Secondary({self.secondary.__class__.__name__})에서 DataFrame 로드 시도: {path}")
        return self.secondary.load_dataframe(path, sheet_name, **kwargs)

    def ensure_directory(self, path: str) -> bool:
        """둘 다 생성 시도합니다.
        
        Args:
            path (str): 생성할 디렉토리 경로.
            
        Returns:
            bool: 성공 여부 (하나라도 성공하면 True).
        """
        r1 = self.primary.ensure_directory(path)
        r2 = self.secondary.ensure_directory(path)
        return r1 or r2

    # --- 쓰기 작업은 Primary에만 수행 (Source 용도이므로) ---
    def save_dataframe_excel(self, df: pd.DataFrame, path: str, **kwargs) -> bool:
        return self.primary.save_dataframe_excel(df, path, **kwargs)

    def save_dataframe_csv(self, df: pd.DataFrame, path: str, **kwargs) -> bool:
        return self.primary.save_dataframe_csv(df, path, **kwargs)

    def save_workbook(self, book: openpyxl.Workbook, path: str) -> bool:
        return self.primary.save_workbook(book, path)

    def get_file(self, path: str) -> Optional[bytes]:
        """Primary에서 로드 시도, 실패 시 Secondary에서 로드합니다.
        
        Args:
            path (str): 파일 경로.
            
        Returns:
            Optional[bytes]: 파일 내용, 실패 시 None.
        """
        if self.primary.path_exists(path):
            data = self.primary.get_file(path)
            if data:
                return data
        
        return self.secondary.get_file(path)

    def put_file(self, path: str, data: bytes) -> bool:
        """Primary에만 저장합니다.
        
        Args:
            path (str): 파일 경로.
            data (bytes): 저장할 데이터.
            
        Returns:
            bool: 저장 성공 여부.
        """
        return self.primary.put_file(path, data)
